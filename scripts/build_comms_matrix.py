#!/usr/bin/env python3
"""
build_comms_matrix.py — Generate the customer-facing DOSTO NEU L2/L3 Network
Communications Matrix as a styled multi-tab Excel workbook.

Sources (design freeze + repo):
  - design freeze/ND-DEL-OBB-035-SDD-002-01_v2.2.docx   (architecture, VLANs, firewall policy, routing)
  - design freeze/network-planning-v7.xlsx               (per-VLAN/per-protocol firewall matrix, subnets)
  - OBN templates nv6/nv4 (per-port switch config — authoritative for 4736/4734)
  - train-ip-allocation-commission/extracted/_shared/{nv6,nv4}-topology.md
  - Stadler IP-Port-Allocation PDFs (per-device authority for 4705/4706)
  - CLAUDE.md playbook (CCU vlan interfaces, vlan7 bit-packed formula)

Output: reports/customer/DOSTO_NEU_Network_Communications_Matrix_v1.0.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ----------------------------------------------------------------------------
# Brand palette (Nomad Digital house style, derived from existing report gens)
# ----------------------------------------------------------------------------
NAVY      = "1F3864"   # title bar
TEAL      = "2E6171"   # section headers
HEADER    = "2E6171"   # table header fill
HEADER_FG = "FFFFFF"
SUBHEAD   = "D6E0E5"   # light teal subheader / banding
BAND      = "F2F6F8"   # zebra band
GREY_TXT  = "595959"
RED       = "C00000"
GREEN     = "548235"
AMBER     = "BF8F00"
STADLER   = "FBE5D6"   # Stadler-owned VLAN row tint (warm)
NOMAD     = "E2F0D9"   # Nomad-owned VLAN row tint (green)
SHARED    = "FFF2CC"   # shared (vlan7) tint (amber)
BORDER_C  = "BFBFBF"

thin = Side(style="thin", color=BORDER_C)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

F_TITLE   = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
F_SUB     = Font(name="Calibri", size=11, color="FFFFFF")
F_SECTION = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
F_HEAD    = Font(name="Calibri", size=10, bold=True, color=HEADER_FG)
F_CELL    = Font(name="Calibri", size=10, color="262626")
F_CELLB   = Font(name="Calibri", size=10, bold=True, color="262626")
F_NOTE    = Font(name="Calibri", size=9, italic=True, color=GREY_TXT)
F_MONO    = Font(name="Consolas", size=9, color="262626")

WRAP_TOP  = Alignment(horizontal="left",   vertical="top", wrap_text=True)
WRAP_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_CTR  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
CTR       = Alignment(horizontal="center", vertical="center")

wb = openpyxl.Workbook()

# ----------------------------------------------------------------------------
# helpers
# ----------------------------------------------------------------------------
def title_bar(ws, text, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, text); c.font = F_TITLE; c.alignment = LEFT_CTR
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    for col in range(1, ncols + 1):
        ws.cell(1, col).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, sub); c.font = F_SUB; c.alignment = LEFT_CTR
    for col in range(1, ncols + 1):
        ws.cell(2, col).fill = PatternFill("solid", fgColor=TEAL)
    ws.row_dimensions[2].height = 22

def section(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, text and 1 or 1, text); c.font = F_SECTION; c.alignment = LEFT_CTR
    for col in range(1, ncols + 1):
        ws.cell(row, col).fill = PatternFill("solid", fgColor=TEAL)
    ws.row_dimensions[row].height = 22
    return row + 1

def header_row(ws, row, headers, widths=None):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row, j, h); c.font = F_HEAD; c.alignment = WRAP_CTR
        c.fill = PatternFill("solid", fgColor=HEADER); c.border = box
    ws.row_dimensions[row].height = 30
    if widths:
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    return row + 1

def data_rows(ws, row, rows, band=True, mono_cols=(), fills=None, valign_top=True):
    for i, r in enumerate(rows):
        fill = None
        if fills and i < len(fills) and fills[i]:
            fill = fills[i]
        elif band and i % 2 == 1:
            fill = BAND
        for j, val in enumerate(r, start=1):
            c = ws.cell(row, j, val)
            c.font = F_MONO if (j in mono_cols) else F_CELL
            c.alignment = WRAP_TOP if valign_top else LEFT_CTR
            c.border = box
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
        row += 1
    return row

def note(ws, row, text, ncols, height=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text); c.font = F_NOTE; c.alignment = WRAP_TOP
    if height:
        ws.row_dimensions[row].height = height
    return row + 1

def blank(ws, row, n=1):
    return row + n

# ============================================================================
# TAB 0 — Cover / Read-me
# ============================================================================
ws = wb.active
ws.title = "0. Overview"
ws.sheet_view.showGridLines = False
NC = 2
title_bar(ws, "DOSTO NEU — Network L2/L3 Communications Matrix",
          "ÖBB Doppelstock-Flotte  ·  Internet-on-Board (Nomad Digital / Stadler)", NC)
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 95
r = 4
meta = [
    ("Document", "Network L2 & L3 Communications Matrix"),
    ("Project", "ND-DEL-OBB-035  ·  DOSTO NEU (Stadler KISS Doppelstock)"),
    ("Customer", "ÖBB — Österreichische Bundesbahnen"),
    ("IoB supplier", "Nomad Digital"),
    ("Vehicle manufacturer", "Stadler Rail"),
    ("Version", "1.1  (firewall tab validated against live CCU iptables, 2026-06-01)"),
    ("Status", "For customer review"),
    ("Fleet scope", "4734 (NV 4-car), 4736 (NV 6-car), 4705 (CAT 5-car), 4706 (FV 6-car)"),
]
for k, v in meta:
    a = ws.cell(r, 1, k); a.font = F_CELLB; a.alignment = LEFT_CTR; a.fill = PatternFill("solid", fgColor=SUBHEAD); a.border = box
    b = ws.cell(r, 2, v); b.font = F_CELL;  b.alignment = LEFT_CTR; b.border = box
    r += 1
r = blank(ws, r)
r = section(ws, r, "Scope & how to read this workbook", NC)
intro = (
    "This workbook documents the Layer-2 (switching) and Layer-3 (routing & firewall) communications design "
    "of the DOSTO NEU on-board network. The CCU (Central Control Unit / Nomad Connect) is the single router, "
    "DHCP/DNS/NTP server and firewall for all Nomad VLANs; Stadler-owned VLANs are routed by the Stadler "
    "Security-Gateway and only transit the Nomad switches as tagged trunk traffic.\n\n"
    "Sources: System Design Document ND-DEL-OBB-035-SDD-002-01 v2.2 and the ÖBB network-planning workbook v0.7 "
    "(design freeze), the Nomad OBN switch-configuration templates (nv6/nv4), and the per-train Stadler "
    "IP-Port-Allocation plans. Where a value is per-train it is given as a formula plus one worked example "
    "(Fzg 133 / 4736-105); per-device IP assignments live in each train's IP-Port-Allocation plan."
)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
c = ws.cell(r, 1, intro); c.font = F_CELL; c.alignment = WRAP_TOP; ws.row_dimensions[r].height = 150
r += 1
r = blank(ws, r)
r = section(ws, r, "Tabs", NC)
tabs = [
    ("1. VLAN Inventory", "All VLANs: ID, name, purpose, owner (Nomad/Stadler), L3 gateway, addressing, services."),
    ("2. CCU VLAN Interfaces", "Per-VLAN CCU L3 interface: role, gateway, subnet, DHCP/DNS/NTP, IP formula + worked example."),
    ("3. L3 Firewall Matrix", "CCU iptables policy: per-VLAN INPUT/FORWARD verdict for each service/protocol (default DENY ALL)."),
    ("4. Stadler Routed Connections", "Explicit wayside connections permitted for Stadler networks via VLAN 7 (the only Stadler→Internet path)."),
    ("5. L2 Switch Port Roles", "Per consist family: the role of each switch port-type (backbone, AP, Stadler-FW, OBS/RDC, access) + STP/PoE policy."),
    ("6. Access Points", "AP→switch/port mapping, carried VLANs, SSID set, authentication."),
    ("7. Consist Topology", "Per family: coach order, switch/AP counts, ring & coupler trunks, Stadler-facing trunk ports."),
    ("8. Addressing Scheme", "The bit-packed DOSTO-NEU IP formula (Stadler VLANs) + Nomad VLAN ranges + backend/MAR allocation."),
]
r = header_row(ws, r, ["Tab", "Contents"], [30, 95])
r = data_rows(ws, r, tabs)
r = blank(ws, r)
note(ws, r, "Nomad Digital — ND-DEL-OBB-035 · DOSTO NEU · Network Communications Matrix v1.0 · generated from design-freeze sources.", NC)

# ============================================================================
# TAB 1 — VLAN Inventory
# ============================================================================
ws = wb.create_sheet("1. VLAN Inventory")
ws.sheet_view.showGridLines = False
NC = 7
title_bar(ws, "VLAN Inventory", "All on-board VLANs — purpose, owner, L3 routing, services", NC)
r = 4
headers = ["VLAN ID", "Name (switch / config)", "Purpose", "Owner / Router",
           "L3 addressing", "DHCP / DNS / NTP", "Services available on VLAN"]
widths  = [9, 24, 30, 18, 30, 22, 46]
r = header_row(ws, r, headers, widths)

# (vid, name, purpose, owner, l3, dhcpdnsntp, services, tint)
vlans = [
    ("1",  "(deactivated)", "Switch-local default VLAN — used only for switch mgmt bootstrap (vlan1 192.168.1.x/24 static on each switch)", "Nomad", "192.168.1.0/24 (switch-local, not routed)", "—", "Switch local management only; not an end-user VLAN", NOMAD),
    ("2",  "zrp-train-repair", "ZFR / FIS central unit (Zugfunkrechner)", "Stadler (Security-Gateway)", "Stadler IP-Port-Allocation plan", "local switch (port-based)", "Stadler-internal; transits Nomad switches as trunk only", STADLER),
    ("3",  "pis-train-repair", "FIS — passenger information displays & speakers", "Stadler", "Stadler IP-Port-Allocation plan", "local switch (port-based)", "Stadler-internal; trunk transport only", STADLER),
    ("5",  "cctv-net", "Video / CCTV (interior + exterior cameras, recorders)", "Stadler", "Stadler IP-Port-Allocation plan", "local switch (port-based)", "Stadler-internal; also the only VLAN bridged over the front coupler (multi-traction)", STADLER),
    ("6",  "reserve-net", "Reservation system / seat displays", "Stadler", "Stadler IP-Port-Allocation plan", "local switch (port-based)", "Stadler-internal; trunk transport only", STADLER),
    ("7",  "obs-net", "Fremdnetz / OBS transit — Stadler↔CCU↔Internet, NTP, DNS", "Stadler ↔ Nomad (shared)", "172.19.x.x/17 (bit-packed; see Addressing)", "n/a (pure transit VLAN)", "NTP (OBS as master, SNAT-forwarded by Stadler FW), routed Internet for explicitly-allowed Stadler connections (Tab 4)", SHARED),
    ("8",  "apc-net", "AFZ — automatic passenger counting", "Stadler", "Stadler IP-Port-Allocation plan", "local switch (port-based)", "Stadler-internal; trunk transport only", STADLER),
    ("9",  "call-point-net", "Sprechstellen — emergency/intercom & ADU", "Stadler", "Stadler IP-Port-Allocation plan", "local switch (port-based)", "Stadler-internal; trunk transport only", STADLER),
    ("12", "call-point-net (energy)", "Energiemessung — energy metering", "Stadler", "Stadler IP-Port-Allocation plan", "local switch (port-based)", "Stadler-internal; isolated, no external interfaces", STADLER),
    ("15", "multitraction-transit", "Multi-traction transit VLAN (coupler)", "Stadler / Nomad", "transit only", "—", "Carried with VLAN 5 over the front-coupler trunk between coupled vehicles", SHARED),
    ("10", "client-net1", "Passenger Wi-Fi — 1st class (\"OEBB\")", "Nomad (CCU)", "192.168.208.0/22  (live; SDD also cites /20)", "DHCP+DNS: CCU", "Captive portal (HTTP/HTTPS, auto-login 24h), Internet, content filter; client isolation", NOMAD),
    ("20", "client-net2", "Passenger Wi-Fi — 2nd class (\"OEBB\")", "Nomad (CCU)", "192.168.224.0/20", "DHCP+DNS: CCU", "Captive portal, Internet, content filter; client isolation", NOMAD),
    ("21-24", "link-sharing-net1..4", "Link-sharing between CCUs (modem/station-WiFi sharing)", "Nomad (CCU)", "10.178.{train}.{link}/30", "—", "Inter-CCU modem pooling; not an end-user VLAN", NOMAD),
    ("30", "staff-net", "On-board staff network (\"iobstaff\")", "Nomad (CCU)", "10.205.{train}.0/24", "DHCP+DNS: CCU", "SSH to CCU/switch/AP, Engineering Pages (https://engineering.page:8005), NMS via VPN, GPS feed; WPA2-Ent; hidden", NOMAD),
    ("31", "zub-priv-nomad-net", "ZuB privileged (conductor devices)", "Nomad (CCU)", "10.204.{train}.0/24", "DHCP+DNS: CCU", "Access to defined ÖBB backend services; WPA2-Ent cert; strong crypto; hidden", NOMAD),
    ("131","zub-none-priv-nomad-net", "ZuB non-privileged (\"OEBBpartner\")", "Nomad (CCU)", "10.211.{train}.0/24", "DHCP+DNS: CCU", "Restricted Internet / defined IPs & ports; WPA2-Ent cert; hidden", NOMAD),
    ("46", "automate-1", "Verpflegungsautomat — status/telemetry", "Nomad (CCU)", "192.168.46.0/24", "DHCP: CCU", "Vending status/telemetry to CDC", NOMAD),
    ("47", "(screen) automate", "Bildschirm Verpflegungsautomat (vending display)", "Nomad (CCU)", "192.168.47.0/24", "DHCP: CCU", "Vending. ACTIVE vending VLAN on NV-4/NV-6/FV-6 families (live VENI/VENF rules). NOT instantiated on CAT 5-car (4705)", NOMAD),
    ("48", "automate-2", "Verpflegungsautomat — payment", "Nomad (CCU)", "192.168.48.0/24", "DHCP: CCU", "Vending payment. On CAT 5-car (4705) this is the ACTIVE vending VLAN (VENI/VENF bound here); on NV/FV-6 it is an unused SVI (46+47 are active). Family-dependent — confirmed live 2026-06-01", NOMAD),
    ("32", "(diagnostics)", "Diagnostics VLAN (live-only — not in design-freeze VLAN list)", "Nomad (CCU)", "(per-train)", "DHCP: CCU", "gpsd 2947, ICMP, NTP 123, DNS 53, Engineering Page 8005 (live DIAGNOSTICI/F chains)", NOMAD),
    ("90", "funksensor-net", "Allgemeine Funksensoren (radio sensors)", "Nomad (reserved)", "10.251.2.0/24", "DHCP: CCU (tbd)", "Reserved / future use; access ports admin-disabled in current templates", NOMAD),
    ("100","mng-nomad-net", "Management VLAN — CCU, switches, APs (OBS-Management)", "Nomad (CCU)", "10.{128+proj}.{train}.128/25  (e.g. 10.179.x.129)", "DHCP+DNS+NTP: CCU", "SNMPv3, SSH, OBN discovery (LLDP), Telemetry/MQTT to backend; native+tagged on trunks", NOMAD),
    ("150","catering-net", "Catering network (\"OEBBcaterer\")", "Nomad (CCU)", "10.251.1.0/24", "DHCP+DNS: CCU", "Pay/order services; WPA2-Ent user; hidden", NOMAD),
    ("200","rdc-nomad-net", "RDC Management (Rail Data Centre)", "Nomad (CCU)", "10.200.0.0/24", "DHCP+DNS: CCU", "Management access to RDC", NOMAD),
    ("202","rdc-interconnect-net", "RDC Interconnect (RDC ↔ TCMS/backend)", "Nomad (CCU)", "10.202.0.0/24", "—", "Data exchange RDC↔backend (NAT via CCU)", NOMAD),
]
fills = [v[7] for v in vlans]
rows  = [v[:7] for v in vlans]
r = data_rows(ws, r, rows, band=False, fills=fills)
for rr in range(5, r):
    ws.row_dimensions[rr].height = 42
r = blank(ws, r)
r = note(ws, r, "Legend:  green = Nomad-owned VLAN (routed by CCU)   ·   orange = Stadler-owned VLAN (routed by Stadler Security-Gateway, trunk-only on Nomad switches)   ·   amber = shared/transit (VLAN 7 / 15).", NC, 28)
r = note(ws, r, "Stadler VLANs (2,3,5,6,8,9,12) are not seen by the CCU at L3 — only the VLAN 7 transit link is. The CCU does not route between Stadler VLANs. Per-device IPs are in each train's IP-Port-Allocation plan; addressing formula on Tab 8.", NC, 40)
ws.freeze_panes = "A5"

# ============================================================================
# TAB 2 — CCU VLAN Interfaces
# ============================================================================
ws = wb.create_sheet("2. CCU VLAN Interfaces")
ws.sheet_view.showGridLines = False
NC = 7
title_bar(ws, "CCU VLAN Interfaces (Layer-3)",
          "The CCU is the router / DHCP / DNS / NTP server / firewall for all Nomad VLANs", NC)
r = 4
r = note(ws, r,
    "The CCU (Nomad Connect, R5001C on the master coach 500/F) terminates one sub-interface per Nomad VLAN and acts as the default gateway. "
    "It runs a strict default-DROP firewall (Tab 3). There is no direct Internet at the CCU — all external traffic is tunnelled via MAR5 VPN to the Nomad backend (NextLayer, Vienna) and only then, if permitted, to the Internet. "
    "Stadler VLANs are NOT terminated on the CCU; only the VLAN 7 transit interface faces the Stadler Security-Gateway.", NC, 56)
r = blank(ws, r)
headers = ["CCU interface", "VLAN", "Role / default gateway", "Subnet (per-train)", "Gateway host", "DHCP / DNS / NTP", "Notes"]
widths  = [16, 8, 26, 30, 16, 22, 40]
r = header_row(ws, r, headers, widths)
ccu_if = [
    ("bond0",   "—",   "WAN uplink — aggregated cellular modems to MAR5 backend", "10.{128+proj}.{train}.1/25 (e.g. 10.179.x.1)", "—", "—", "Not a passenger/Stadler interface; MAR5 VPN endpoint; load-balanced modems"),
    ("vlan100", "100", "Management gateway (CCU, switches, APs)", "10.{128+proj}.{train}.128/25", ".129", "DHCP+DNS+NTP", "Native untagged on switch trunk ports; OBN/LLDP discovery; SNMPv3; MQTT telemetry"),
    ("vlan10",  "10",  "1st-class passenger Wi-Fi gateway", "192.168.208.0/20", ".1", "DHCP+DNS", "Captive portal; Internet via backend; client isolation"),
    ("vlan20",  "20",  "2nd-class passenger Wi-Fi gateway", "192.168.224.0/20", ".1", "DHCP+DNS", "Captive portal; Internet via backend; client isolation"),
    ("vlan30",  "30",  "On-board staff gateway (iobstaff)", "10.205.{train}.0/24", ".1", "DHCP+DNS", "Engineering Pages :8005; SSH to infra; NMS via VPN; GPS/NMEA feed"),
    ("vlan31",  "31",  "ZuB privileged gateway", "10.204.{train}.0/24", ".1", "DHCP+DNS", "Defined ÖBB backend services only"),
    ("vlan131", "131", "ZuB non-privileged gateway", "10.211.{train}.0/24", ".1", "DHCP+DNS", "Restricted Internet"),
    ("vlan46",  "46",  "Vending status/telemetry gateway", "192.168.46.0/24", ".1", "DHCP", "To CDC status service"),
    ("vlan47",  "47",  "Vending display gateway", "192.168.47.0/24", ".1", "DHCP", "Vending screens"),
    ("vlan48",  "48",  "Vending payment gateway", "192.168.48.0/24", ".1", "DHCP", "Restricted; NAT via CCU"),
    ("vlan150", "150", "Catering gateway", "10.251.1.0/24", ".1", "DHCP+DNS", "Pay/order services"),
    ("vlan200", "200", "RDC management gateway", "10.200.0.0/24", ".1", "DHCP+DNS", "Rail Data Centre mgmt"),
    ("vlan202", "202", "RDC interconnect", "10.202.0.0/24", ".1", "—", "RDC↔backend; NAT via CCU"),
    ("vlan7",   "7",   "Stadler transit interface (Fremdnetz)", "172.19.<128+Fzg//2>.<dev>/17", "FW = host .1 of /17", "NTP relay only", "Worked example Fzg 133: CCU 172.19.194.130/17, Stadler FW 172.19.194.129. Even Fzg→host .2, odd Fzg→host .130 (device 2). See Tab 8."),
]
r = data_rows(ws, r, ccu_if, mono_cols=(4,5))
for rr in range(8, r):
    ws.row_dimensions[rr].height = 34
r = blank(ws, r)
r = note(ws, r, "{proj} = project ID (DOSTO NEU backend uses 10.178.0.0/17 and 10.179.0.0/17; on-board management lands in 10.179.x.x). {train} = train ID. Per-train gateway IPs and DHCP pools are defined in each train's IP-Port-Allocation plan.", NC, 40)
r = note(ws, r, "vlan1 (192.168.1.0/24) exists only switch-locally for bootstrap and is not routed by the CCU. VLAN 90 (Funksensor) is reserved; its access ports are admin-disabled in the current OBN templates.", NC, 28)
ws.freeze_panes = "A8"

# ============================================================================
# TAB 3 — L3 Firewall Matrix
# ============================================================================
ws = wb.create_sheet("3. L3 Firewall Matrix")
ws.sheet_view.showGridLines = False
svc = ["ICMP", "SSH\n22/tcp", "DNS\n53", "DHCP\n67/udp", "HTTP\n80/tcp", "NTP\n123/udp", "HTTPS\n443/tcp", "MQTT/S\n1883/8883", "GPS\n2947/tcp", "Eng.Page\n8005/tcp"]
NC = 3 + len(svc)*2 + 1   # name, vlan, owner + (INPUT/FWD)*svc + other
title_bar(ws, "Layer-3 Firewall Matrix — CCU (I²OB Router)",
          "Default policy: DROP on INPUT + FORWARD; ACCEPT on OUTPUT. FLEET-COMMON across all 4 families — validated live on nv6/4736 (Fzg 146,130), nv4/4734 (Fzg 11), FV6/4706 (Fzg 190), CAT5/4705 (Fzg 229), 2026-06-01.", NC)
r = 4
r = note(ws, r,
    "INPUT = traffic destined to the CCU itself (e.g. DHCP/DNS/NTP server, portal).   FORWARD = traffic routed through the CCU (e.g. to the Internet via backend).   "
    "Cells: allow / deny / —(n/a or undefined).   Passenger VLANs may never reach RFC1918 private destinations (all blocked). All passenger clients are isolated from each other.   "
    "Stadler VLANs (2,3,5,6,8,9,12) are not routed by the CCU at all — they have no CCU firewall entry; their only external path is VLAN 7 (Tab 4).", NC, 56)
r = blank(ws, r)

# two-row header: service groups, then INPUT/FORWARD
top = r
ws.cell(top, 1, "Network"); ws.cell(top, 2, "VLAN"); ws.cell(top, 3, "Owner")
for k in (1,2,3):
    ws.merge_cells(start_row=top, start_column=k, end_row=top+1, end_column=k)
    c = ws.cell(top, k); c.font = F_HEAD; c.alignment = WRAP_CTR; c.fill = PatternFill("solid", fgColor=HEADER); c.border = box
col = 4
for s in svc:
    ws.merge_cells(start_row=top, start_column=col, end_row=top, end_column=col+1)
    c = ws.cell(top, col, s); c.font = F_HEAD; c.alignment = WRAP_CTR; c.fill = PatternFill("solid", fgColor=HEADER); c.border = box
    ws.cell(top, col+1).fill = PatternFill("solid", fgColor=HEADER); ws.cell(top, col+1).border = box
    col += 2
ws.merge_cells(start_row=top, start_column=col, end_row=top+1, end_column=col)
c = ws.cell(top, col, "Other allowed"); c.font = F_HEAD; c.alignment = WRAP_CTR; c.fill = PatternFill("solid", fgColor=HEADER); c.border = box
# sub-header row
sub = top + 1
col = 4
for s in svc:
    for lbl in ("IN", "FWD"):
        c = ws.cell(sub, col, lbl); c.font = F_HEAD; c.alignment = WRAP_CTR; c.fill = PatternFill("solid", fgColor=TEAL); c.border = box
        col += 1
ws.row_dimensions[top].height = 30
ws.row_dimensions[sub].height = 16
# widths
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 7
ws.column_dimensions["C"].width = 9
for j in range(4, 4+len(svc)*2):
    ws.column_dimensions[get_column_letter(j)].width = 6
ws.column_dimensions[get_column_letter(col)].width = 34

# rows: each = name, vlan, owner, then 10*(IN,FWD), then other
# order: ICMP, SSH, DNS, DHCP, HTTP, NTP, HTTPS, MQTT, GPS, EngPage
def fwrow(name, vlan, owner, cells, other):
    return [name, vlan, owner] + cells + [other]

A = "allow"; D = "deny"; N = "—"
fw = [
    fwrow("VLAN 1 (switch default)", "1", "Nomad",
          [D,A, D,A, A,A, A,D, A,A, A,A, A,A, D,A, D,N, D,N], "Switch-local; not an end-user path"),
    fwrow("1st-class Wi-Fi", "10", "Nomad",
          [D,A, D,A, A,A, A,D, A,A, A,A, A,A, D,A, D,N, D,N], "Internet via backend; no RFC1918"),
    fwrow("2nd-class Wi-Fi", "20", "Nomad",
          [D,A, D,A, A,A, A,D, A,A, A,A, A,A, D,A, D,N, D,N], "Internet via backend; no RFC1918"),
    fwrow("On-board staff (iobstaff)", "30", "Nomad",
          [A,N, A,N, A,A, A,D, A,N, A,N, A,N, A,N, A,N, A,N], "Access to router/mgmt; Eng.Pages; NMS via VPN"),
    fwrow("ZuB privileged", "31", "Nomad",
          [A,A, D,N, A,A, A,D, A,N, A,N, A,N, N,N, A,N, A,N], "Defined ÖBB IPs & ports only (strong crypto)"),
    fwrow("ZuB non-privileged", "131", "Nomad",
          [A,A, D,N, A,A, A,D, A,N, A,N, A,N, N,N, A,N, A,N], "Restricted Internet / defined IPs & ports"),
    fwrow("Catering", "150", "Nomad",
          [A,N, D,N, A,A, A,D, D,N, A,N, D,N, N,N, D,N, D,N], "Pay/order services (restricted)"),
    fwrow("Vending status", "46", "Nomad",
          [A,N, D,N, A,N, A,D, D,N, A,N, D,N, N,N, N,N, N,N], "To CDC status service"),
    fwrow("Vending display", "47", "Nomad",
          [A,N, D,N, A,N, A,D, D,N, A,N, D,N, N,N, N,N, N,N], "Vending screens"),
    fwrow("Vending payment", "48", "Nomad",
          [A,N, D,N, A,N, A,D, D,N, A,N, D,N, N,N, N,N, N,N], "Payment (restricted, NAT via CCU)"),
    fwrow("Management", "100", "Nomad",
          [A,A, A,A, A,A, A,D, A,A, A,A, A,A, A,A, A,A, A,A], "SNMPv3, SSH, MQTT telemetry, OBN/LLDP"),
    fwrow("RDC management", "200", "Nomad",
          [A,N, A,N, A,N, A,D, A,N, A,N, A,N, N,N, N,N, N,N], "RDC mgmt"),
    fwrow("RDC interconnect", "202", "Nomad",
          [A,N, N,N, N,N, N,N, N,N, N,N, N,N, A,A, N,N, N,N], "RDC↔backend; NAT via CCU"),
    fwrow("Funksensor", "90", "Nomad",
          [N,N, N,N, N,N, N,N, N,N, N,N, N,N, N,N, N,N, N,N], "Reserved — no rules defined yet"),
    fwrow("Link sharing", "21-24", "Nomad",
          [N,N, N,N, N,N, N,N, N,N, N,N, N,N, N,N, N,N, N,N], "Inter-CCU only (tbd); not end-user"),
    fwrow("Fremdnetz / OBS transit", "7", "Stadler↔Nomad",
          [N,N, N,N, A,N, N,N, N,N, A,N, N,N, N,N, N,N, N,N], "Default-deny + explicit allow; only NTP/DNS to CCU; routed Internet only for allowed Stadler dest (Tab 4); RFC1918 dest dropped"),
]
rr = sub + 1
for i, row in enumerate(fw):
    owner_tint = STADLER if "Stadler" in row[2] else None
    fill = owner_tint or (BAND if i % 2 == 1 else None)
    for j, val in enumerate(row, start=1):
        c = ws.cell(rr, j, val)
        c.border = box
        if j <= 3:
            c.font = F_CELLB if j == 1 else F_CELL
            c.alignment = LEFT_CTR if j == 1 else CTR
            c.fill = PatternFill("solid", fgColor=fill) if fill else PatternFill("solid", fgColor=SUBHEAD)
        elif j == len(row):
            c.font = F_CELL; c.alignment = WRAP_TOP
            if fill: c.fill = PatternFill("solid", fgColor=fill)
        else:
            c.font = F_CELL; c.alignment = CTR
            if val == A:
                c.fill = PatternFill("solid", fgColor=NOMAD); c.font = Font(name="Calibri", size=9, bold=True, color=GREEN)
            elif val == D:
                c.fill = PatternFill("solid", fgColor="FCE4E4"); c.font = Font(name="Calibri", size=9, color=RED)
            else:
                c.font = Font(name="Calibri", size=9, color="A6A6A6")
                if fill and fill != BAND: c.fill = PatternFill("solid", fgColor=fill)
    ws.row_dimensions[rr].height = 30
    rr += 1
r = rr + 1
r = note(ws, r, "Verdicts cross-checked against the live CCU ruleset (nf_tables, iptables-save) on Fzg 146 & 130, 2026-06-01 — and against the ÖBB network-planning workbook v0.7 'Firewall' tab + SDD §6.13. Blank cells = no explicit rule (default DROP applies). Live zone chains: PASSENGER_INPUT/FORWARD (vlan10), MGMTI/F (vlan100), FISI/F (vlan7), PDA1I/F (vlan30 staff), VENI/F (vlan46/47), DIAGNOSTICI/F (vlan32), MAR3TUNI/F (mar5-tun), LAN2I (wired svc LAN).", NC, 40)
r = note(ws, r, "Live-confirmed specifics: (1) RFC1918 is dropped before every Internet egress (ce0-3p0) and inside PASSENGER/MGMT/FIS/PDA/VEN forward chains. (2) ICMP-to-CCU is allowed from management, staff(vlan30), diagnostics(vlan32) — not from passengers. (3) MQTT 1883/8883, GPS 1816/2947, MA-API 6001/6002/8443 are CCU-internal Docker services reached via the MAR5 tunnel. (4) QoS via TOS marking: management DSCP 0xe0, staff 0xc0, gold/cf1 0x40, silver/cf2 0x20.", NC, 44)
r = note(ws, r, "Per-train caveat: this matrix is the fleet DESIGN. A given train instantiates only the VLAN interfaces it needs (e.g. Fzg 146/130/11/190/229 have vlan 7,10,30,46,100 + one of 47/48; none had vlan20/31/131/150/200/202/90 live). The core ruleset is Puppet-driven and identical across all families.", NC, 30)
r = note(ws, r, "Family-specific delta found live (2026-06-01): CAT 5-car (4705) binds the vending firewall zone to vlan48; NV-4/NV-6/FV-6 use vlan47. Otherwise the VLAN-7 transit posture is the FISF RFC1918-drop chain on ALL families (the design baseline shown in this matrix). NB: 2 of 6 sampled CCUs (Fzg 11,12) carried stray RUNTIME-only 'FORWARD vlan7 ACCEPT' rules above FISF — a manually-injected debug leftover (not persisted, not design); recommend flush/reboot on those boxes. Matrix reflects the intended posture.", NC, 44)
ws.freeze_panes = "D" + str(sub+1)

# ============================================================================
# TAB 4 — Stadler Routed Connections (VLAN 7)
# ============================================================================
ws = wb.create_sheet("4. Stadler Routed Conns")
ws.sheet_view.showGridLines = False
NC = 7
title_bar(ws, "Permitted Stadler Wayside Connections (via VLAN 7)",
          "The only path from Stadler networks to the wayside — NAT'd by the CCU after firewall inspection", NC)
r = 4
r = note(ws, r,
    "Stadler-owned VLANs reach the wayside only through VLAN 7 (Fremdnetz/OBS transit). The Stadler Security-Gateway NATs its traffic onto VLAN 7; the CCU inspects it against a default-deny policy and forwards permitted flows via the MAR5 backend. "
    "Private (RFC1918) destinations are dropped. The connections below are the explicitly-agreed wayside endpoints (SDD §6.18.1 'Erlaubte Verbindungen für Stadler-Netzwerke').", NC, 44)
r = blank(ws, r)
headers = ["Description", "Direction", "Source", "Destination", "Port", "Protocol", "Notes"]
widths  = [34, 12, 18, 22, 9, 11, 30]
r = header_row(ws, r, headers, widths)
stad = [
    ("VPN tunnel to RDS", "Train → wayside", "Stadler Router", "62.2.130.53", "83", "UDP", "Stadler RDS VPN"),
    ("VPN tunnel to RDS", "Train → wayside", "Stadler Router", "62.2.130.53", "84", "UDP", "Stadler RDS VPN (2nd port)"),
    ("Data transfer to ÖBB landside", "Train → wayside", "Stadler Router", "195.69.192.153", "22", "TCP", "SFTP/SSH to ÖBB landside"),
    ("IPSec tunnel to video server", "Train → wayside", "Stadler Router", "TBD", "500", "UDP", "IKE/IPSec; destination to be defined by Stadler"),
]
r = data_rows(ws, r, stad, mono_cols=(4,5))
for rr in range(7, r):
    ws.row_dimensions[rr].height = 24
r = blank(ws, r)
r = section(ws, r, "Time synchronisation across the VLAN-7 boundary (NTP)", NC)
r = note(ws, r,
    "The OBS acts as the on-board NTP master (primary source: landside Internet NTP; fallback: GPS). NTP is published into VLAN 7 so non-IoB (Stadler) devices can take OBS time. "
    "The Stadler-side RCU runs its own NTP server (e.g. 192.168.101.10/.15, port-based) that uses the OBS as authoritative reference. NTP requests from the Stadler side are directed to the Stadler firewall, which forwards them to the OBS and applies SNAT so replies return correctly. "
    "All Stadler subsystems take time only from their own NTP server — logical VLAN isolation is preserved despite shared switch hardware.", NC, 56)
r = blank(ws, r)
r = note(ws, r, "Routing responsibility for VLAN 7 is shared: Stadler operates the Security-Gateway that performs inter-VLAN routing for all Stadler device VLANs; Nomad's CCU provides the firewalled wayside/Internet path. Multi-traction routing of Stadler services over the coupler is a Variation Order (ND-036 ch.6), not in the standard scope.", NC, 40)
ws.freeze_panes = "A7"

# ============================================================================
# TAB 5 — L2 Switch Port Roles
# ============================================================================
ws = wb.create_sheet("5. L2 Switch Port Roles")
ws.sheet_view.showGridLines = False
NC = 6
title_bar(ws, "Layer-2 Switch Port Roles",
          "VDSRail DS-CSWP28-10GbE — port-role taxonomy across all consist families", NC)
r = 4
r = note(ws, r,
    "Each coach carries VDSRail 28-port consist switches (3 per coach; the catering/PRM coaches carry a 4th). "
    "Port roles below are common across all four families; the exact port NUMBER a role lands on is Stadler-defined per consist and listed in each train's IP-Port-Allocation plan. "
    "The nv6 (4736) and nv4 (4734) per-port assignments are fixed by the Nomad OBN templates; 4705 (CAT) and 4706 (FV) follow the same role set with family-specific port positions.", NC, 48)
r = blank(ws, r)
headers = ["Port role", "Typical port(s)", "Switchport mode", "VLAN membership", "STP / PoE", "Notes"]
widths  = [22, 18, 16, 34, 20, 40]
r = header_row(ws, r, headers, widths)
roles = [
    ("Backbone ring uplink", "e0-0, e0-1", "Trunk (all VLANs, prune 1-1000)", "All VLANs (1-1000)", "edge OFF; Coldrelay-bypass", "Forms the per-vehicle RSTP ring. e0-0 = intra-coach to sibling switch; e0-1 = inter-coach to next coach. 10 GbE. Hardware-bridged on switch failure."),
    ("Front-coupler trunk", "e0-2 (A1, A3, B1, B3 only)", "Trunk (prune 5,15)", "VLAN 5 (Video) + 15 (multi-traction transit)", "edge OFF; high path-cost", "Connects to next coupled vehicle (Dellner OctiConn, 1 GbE). DOWN when solo. Path-cost = train_id×2,000,000 (A3) / ×1,000,000 (A1) so internal ring is preferred."),
    ("Stadler firewall trunk", "e1-4 (A3 only)", "Trunk (prune 2,3,5,6,7,8,9,12,15)", "All Stadler VLANs + VLAN 7 transit", "edge ON; PoE critical", "Single uplink to the Stadler Security-Gateway. Carries every Stadler device VLAN plus the VLAN-7 transit. 1 GbE."),
    ("OBS trunk", "e0-2 (D1/D3 — nv6; G1/G3 — nv4)", "Trunk (Nomad VLANs + 7)", "100,10,20,21-24,30,31,46,48,90,131,150,200,202,7", "edge ON", "Uplink to the On-Board-System. 10 GbE. D2/G2 OBS port is a disabled reserve."),
    ("RDC trunk", "e0-3 (D1/D3 — nv6; G1/G3 — nv4)", "Trunk (prune 200,202)", "VLAN 200 + 202", "edge ON", "Rail Data Centre uplink. 10 GbE; often idle if RDC powered off. D2/G2 RDC port is a disabled reserve."),
    ("ZFR access", "e1-11 (B1 primary, B3 standby)", "Access VLAN 2", "VLAN 2 only", "edge ON", "Zugfunkrechner. B1/B3 are a redundant pair sharing one IP — only one transmits at a time."),
    ("Access-Point trunk", "e0-4 (each switch) + e1-2 (3rd switch)", "Trunk (prune 100,10,20,30,31,131,150,1)", "Mgmt + all passenger/staff VLANs", "edge ON; PoE critical", "Westermo Ibex-610 APs. 4 APs per coach (AP1-3 on e0-4 of the 3 switches; AP4 on e1-2 of the 3rd switch). 1 GbE."),
    ("Camera access (CCTV)", "various e1-x / e0-x", "Access VLAN 5", "VLAN 5", "edge ON; PoE critical", "Interior/exterior cameras + recorders. Port-based DHCP from local switch (video group). Count varies per coach."),
    ("FIS / display access", "various e1-x / e2-x", "Access VLAN 3", "VLAN 3", "edge ON", "Side displays, front displays, advertising screens. Port-based DHCP (fis group)."),
    ("AFZ access", "e1-x", "Access VLAN 8", "VLAN 8", "edge ON; PoE", "Automatic passenger counting sensors. Port-based DHCP (afz group)."),
    ("Intercom / Sprechstelle access", "e1-x / e2-x", "Access VLAN 9", "VLAN 9", "edge ON; PoE critical", "Emergency call points + ADU + audio amps. Port-based DHCP (sprechstelle group). PRM-accessible variants on nv4 E-coach."),
    ("Energy meter access", "e2-x (E coach on nv4)", "Access VLAN 12", "VLAN 12", "edge ON", "Energiezähler. Isolated VLAN, no external interface."),
    ("Funksensor access", "e1-12, e1-13", "Access VLAN 90", "VLAN 90", "admin-disabled", "Radio-sensor future-use ports; currently 'no enable' in templates."),
    ("Service FIS / CCTV", "e0-x", "Access VLAN 2", "VLAN 2", "edge ON", "Redundancy/service port for FIS-CCTV (seen on 4705/4706 plans)."),
    ("vlan1 (mgmt bootstrap)", "interface vlan1", "L3 SVI", "192.168.1.x/24 (switch-local)", "lldp off", "Each switch has a static vlan1 IP for local bootstrap; not routed."),
    ("vlan100 (mgmt)", "interface vlan100", "L3 SVI (DHCP)", "VLAN 100", "—", "Switch management address via CCU DHCP; SNMPv3 + SSH + NTP client to 10.x.x.1."),
]
r = data_rows(ws, r, roles)
for rr in range(7, r):
    ws.row_dimensions[rr].height = 46
r = blank(ws, r)
r = note(ws, r, "Switch config facts (modes, prune lists, edge/PoE, path-cost) are taken from the Nomad OBN nv6/nv4 templates (e.g. nv6-100-A3.cfg). Switch admin/SNMP: SNMPv3 user 'snmpadmin', community read-write; SSH + HTTPS-REST + TFTP services enabled; Telnet/HTTP-REST disabled.", NC, 36)
r = note(ws, r, "All client-facing ports use 'spanning-tree edge on' so they go straight to forwarding (no STP recompute on plug-in). Backbone ports use 'edge off' and participate in RSTP. Coldrelay-bypass hardware-bridges the two backbone ports if a switch fails — complementary to RSTP (which covers link/cable faults).", NC, 36)
ws.freeze_panes = "A7"

# ============================================================================
# TAB 6 — Access Points
# ============================================================================
ws = wb.create_sheet("6. Access Points")
ws.sheet_view.showGridLines = False
NC = 5
title_bar(ws, "Access Points (Wi-Fi)",
          "Westermo Ibex-610 (NMID 1043) — 4 per coach; AP→switch/port mapping & SSID set", NC)
r = 4
r = note(ws, r,
    "Each coach has 4 APs. On every coach: AP1/AP2/AP3 hang off port e0-4 of the coach's three switches; AP4 hangs off e1-2 of the third switch. "
    "All AP uplinks are trunks carrying VLANs 100,10,20,30,31,131,150,1. Indoor antennas: Huber+Suhner 1399.35.0020.", NC, 40)
r = blank(ws, r)
r = section(ws, r, "AP → switch / port mapping (per coach pattern — applies to every coach of every family)", NC)
headers = ["AP slot", "On switch", "Port", "Uplink mode", "Carried VLANs"]
widths  = [12, 22, 10, 16, 40]
r = header_row(ws, r, headers, widths)
aps = [
    ("AP1", "1st switch of coach (x1)", "e0-4", "Trunk", "100, 10, 20, 30, 31, 131, 150, 1"),
    ("AP2", "2nd switch of coach (x2)", "e0-4", "Trunk", "100, 10, 20, 30, 31, 131, 150, 1"),
    ("AP3", "3rd switch of coach (x3)", "e0-4", "Trunk", "100, 10, 20, 30, 31, 131, 150, 1"),
    ("AP4", "3rd switch of coach (x3)", "e1-2", "Trunk", "100, 10, 20, 30, 31, 131, 150, 1"),
]
r = data_rows(ws, r, aps, mono_cols=(5,))
for rr in range(r-4, r):
    ws.row_dimensions[rr].height = 20
r = blank(ws, r)
r = section(ws, r, "AP count per family", NC)
headers = ["Family", "Consist", "Coaches", "Switches", "APs"]
r = header_row(ws, r, headers, [16, 18, 28, 12, 8])
apcount = [
    ("4734", "NV 4-car (nv4)", "A · G · E · B", "12", "16"),
    ("4736", "NV 6-car (nv6)", "A · C · D · E · F · B", "18", "24"),
    ("4705", "CAT 5-car (fv5)", "A · C · E · F · B (E has 4 sw)", "~16", "~20"),
    ("4706", "FV 6-car (nv6-style)", "A · C · D · E · F · B", "18", "24"),
]
r = data_rows(ws, r, apcount)
for rr in range(r-4, r):
    ws.row_dimensions[rr].height = 20
r = blank(ws, r)
r = section(ws, r, "SSID set (broadcast by the APs)", NC)
headers = ["SSID", "VLAN", "Class / use", "Authentication", "Visible / Landing page"]
r = header_row(ws, r, headers, [16, 8, 26, 22, 26])
ssids = [
    ("OEBB", "10 / 20", "Passenger 1st / 2nd class", "Open", "Visible · captive portal, 12h"),
    ("iobstaff", "30", "On-board staff", "WPA2-Enterprise (cert)", "Hidden · no landing page"),
    ("OEBBzub", "31", "ZuB privileged", "WPA2-Enterprise (cert)", "Hidden"),
    ("OEBBpartner", "131", "ZuB non-privileged", "WPA2-Enterprise (cert)", "Hidden"),
    ("OEBBcaterer", "150", "Catering", "WPA2-Enterprise (user)", "Hidden"),
]
r = data_rows(ws, r, ssids)
for rr in range(r-5, r):
    ws.row_dimensions[rr].height = 20
r = blank(ws, r)
r = note(ws, r, "All passenger/staff SSIDs enforce client (user) separation. Stadler VLANs carry no Wi-Fi. The 'm-' AP config variant (middle coaches) differs only in the AP-side LuCI config, not in the switch trunk membership. Per-AP management IPs are issued by CCU DHCP on VLAN 100 and listed in the IP-Port-Allocation plan.", NC, 36)
ws.freeze_panes = "A8"

# ============================================================================
# TAB 7 — Consist Topology
# ============================================================================
ws = wb.create_sheet("7. Consist Topology")
ws.sheet_view.showGridLines = False
NC = 7
title_bar(ws, "Consist Topology (L1/L2 Backbone)",
          "Per family — coach order, device counts, ring & coupler trunks, Stadler-facing ports", NC)
r = 4
r = note(ws, r,
    "The backbone is a per-vehicle RSTP ring of VDSRail switches. Coach order is front→back. Each coach's switches are ring-linked via e0-0 (intra-coach) and e0-1 (inter-coach). "
    "The two end coaches' switches additionally expose a front-coupler trunk (e0-2) for multi-traction. STP root = the switch on the master coach (priority 0); convergence ~10 s.", NC, 44)
r = blank(ws, r)
headers = ["Family", "Type", "Coach order (front→back)", "Switches", "APs", "Master coach / CCU", "Stadler-facing trunk ports"]
widths  = [12, 18, 26, 12, 8, 22, 34]
r = header_row(ws, r, headers, widths)
topo = [
    ("4734", "NV 4-car (nv4)", "A(100) · G(200) · E(400) · B(600)", "12 (3/coach)", "16",
     "no on-board master*; CCU on master variant", "A3 e1-4 (FW); G1/G3 e0-2 (OBS), e0-3 (RDC); B1/B3 e1-11 (ZFR)"),
    ("4736", "NV 6-car (nv6)", "A(100) · C(200) · D(300) · E(400) · F(500) · B(600)", "18 (3/coach)", "24",
     "F coach (500) — R5001C + 4× roof antenna", "A3 e1-4 (FW); D1/D3 e0-2 (OBS), e0-3 (RDC); B1/B3 e1-11 (ZFR)"),
    ("4705", "CAT 5-car (fv5)", "A · C · E · F · B  (E coach has 4 switches)", "~16", "~20",
     "F coach — CCU R5001C", "A3 e1-4 (FW); OBS/RDC on G/D-equivalent; ZFR on B; + catering VLANs 46/47/48"),
    ("4706", "FV 6-car (nv6-style)", "A(100) · C(200) · D(300) · E(400) · F(500) · B(600)", "18 (3/coach)", "24",
     "F coach (500) — R5001C", "A3 e1-4 (FW); D1/D3 e0-2 (OBS), e0-3 (RDC); B1/B3 e1-11 (ZFR)"),
]
r = data_rows(ws, r, topo)
for rr in range(7, r):
    ws.row_dimensions[rr].height = 50
r = blank(ws, r)
r = section(ws, r, "Backbone ring & redundancy (all families)", NC)
ring = [
    ("Ring topology", "Per-vehicle ring over the coach switches; RSTP (IEEE 802.1w) blocks one path in normal operation."),
    ("Intra-coach link", "e0-0 of each switch → sibling switch in the same coach."),
    ("Inter-coach link", "e0-1 of each switch → a switch in the adjacent coach (10 GbE)."),
    ("STP root election", "Switch on the master coach = priority 0 (root); secondary at same end = 4096; rest = default. Fallback by MAC."),
    ("Convergence", "~10 s RSTP (vs Stadler's 20 ms target — noted as a requirement gap; standard scope)."),
    ("Front coupler", "e0-2 on A1, A3, B1, B3 → next vehicle (Dellner OctiConn, 1 GbE). Carries VLAN 5 + 15. High path-cost = protection path only."),
    ("Coldrelay-bypass", "Hardware bridges a failed switch's two backbone ports so the ring survives a switch hardware failure (complements RSTP)."),
    ("Edge ports", "All client/AP/3rd-party ports = 'spanning-tree edge' → immediate forwarding, no ring recompute on plug-in."),
    ("Multi-traction", "Series/daisy-chain between vehicles (NOT a ring). RSTP transparent over couplers. Routing of Stadler services over couplers = Variation Order (ND-036 ch.6)."),
]
r = header_row(ws, r, ["Aspect", "Design"], [22, 110])
r = data_rows(ws, r, ring)
for rr in range(r-len(ring), r):
    ws.row_dimensions[rr].height = 28
r = blank(ws, r)
r = note(ws, r, "* On the 4-car (nv4), the CCU is fitted on the master variant of the consist; coach order A·G·E·B (G replaces nv6's C/D function). Switch/AP counts and the exact port positions for 4705/4706 are authoritative in each train's Stadler IP-Port-Allocation plan; counts shown for 4705 are approximate pending per-train extraction.", NC, 40)
ws.freeze_panes = "A7"

# ============================================================================
# TAB 8 — Addressing Scheme
# ============================================================================
ws = wb.create_sheet("8. Addressing Scheme")
ws.sheet_view.showGridLines = False
NC = 6
title_bar(ws, "IP Addressing Scheme",
          "Bit-packed DOSTO-NEU scheme (Stadler VLANs) + Nomad VLAN ranges + backend allocation", NC)
r = 4
r = section(ws, r, "Bit-packed address format (Stadler VLANs, 172.16.0.0/12)", NC)
r = note(ws, r,
    "Stadler device VLANs use a 32-bit packed scheme so every IP is fleet-unique. VLANs are carved into /17 subnets. "
    "Per VLAN+port a fixed IP-ID is assigned; the train (Fzg) ID makes the address unique across the fleet.", NC, 32)
headers = ["Bits", "Field", "Range", "Meaning"]
r = header_row(ws, r, headers, [12, 18, 14, 70])
bits = [
    ("1-8",  "Prefix (octet 1)", "172", "Static private-range prefix"),
    ("9-12", "Prefix ext.", "16-31", "Static private-range extension (172.16.0.0/12)"),
    ("13-17","VLAN ID", "1-31", "5-bit VLAN identifier (e.g. VLAN 7 = 00111)"),
    ("18-25","Fahrzeug (Fzg) ID", "1-255", "8-bit vehicle number — numbers all vehicles 1..n regardless of type"),
    ("26-32","Device", "1-127", "7-bit network-device number (CCU on VLAN 7 = device 2; Stadler FW = device 1)"),
]
r = data_rows(ws, r, bits, mono_cols=(3,))
for rr in range(r-len(bits), r):
    ws.row_dimensions[rr].height = 22
r = blank(ws, r)
r = section(ws, r, "CCU vlan7 IP — formula & worked example (Fzg 133 / 4736-105)", NC)
headers = ["Quantity", "Formula", "Fzg 133 (odd)", "Even-Fzg variant"]
r = header_row(ws, r, headers, [22, 40, 26, 26])
v7 = [
    ("octet 3", "128 + (Fzg // 2)", "128 + 66 = 194", "same"),
    ("octet 4 (CCU, device 2)", "(128 if Fzg odd else 0) + 2", "128 + 2 = 130", "0 + 2 = 2"),
    ("octet 4 (Stadler FW, device 1)", "(128 if Fzg odd else 0) + 1", "128 + 1 = 129", "0 + 1 = 1"),
    ("CCU vlan7 address", "172.19.<o3>.<o4>/17", "172.19.194.130/17", "172.19.<o3>.2/17"),
    ("Stadler FW (peer)", "172.19.<o3>.<o4>", "172.19.194.129", "172.19.<o3>.1"),
    ("Subnet", "172.19.128.0/17", "172.19.128.0/17", "172.19.128.0/17"),
]
r = data_rows(ws, r, v7, mono_cols=(3,4))
for rr in range(r-len(v7), r):
    ws.row_dimensions[rr].height = 22
r = blank(ws, r)
r = note(ws, r,
    "IMPORTANT: the formula in /etc/nd-redundancy/networks.yaml on production CCUs computes from OBN's train_id, not Fzg ID, and is wrong. "
    "The live vlan7 IP is set per-train in the NetworkManager connection (ndrd-vlan-vlan7). Verify before any health check — Stadler reachability depends on it.", NC, 36)
r = blank(ws, r)
r = section(ws, r, "Nomad VLAN ranges (operator / passenger / management)", NC)
headers = ["Network", "Subnet", "Notes"]
r = header_row(ws, r, headers, [30, 30, 60])
nom = [
    ("Client 1st class", "192.168.208.0/20", "Passenger Wi-Fi VLAN 10"),
    ("Client 2nd class", "192.168.224.0/20", "Passenger Wi-Fi VLAN 20"),
    ("Management VLAN", "10.{128+proj}.{train}.128/25", "VLAN 100 — CCU/switch/AP; CCU = .129"),
    ("On-board staff", "10.205.{train}.0/24", "VLAN 30"),
    ("ZuB privileged", "10.204.{train}.0/24", "VLAN 31"),
    ("ZuB non-privileged", "10.211.{train}.0/24", "VLAN 131"),
    ("Catering", "10.251.1.0/24", "VLAN 150"),
    ("Funksensor", "10.251.2.0/24", "VLAN 90 (reserved)"),
    ("Link sharing", "10.178.{train}.{link}/30", "VLANs 21-24 (inter-CCU)"),
    ("RDC management", "10.200.0.0/24", "VLAN 200"),
    ("RDC interconnect", "10.202.0.0/24", "VLAN 202"),
    ("Vending status / display / pay", "192.168.46/47/48.0/24", "VLANs 46 / 47 / 48"),
]
r = data_rows(ws, r, nom, mono_cols=(2,))
for rr in range(r-len(nom), r):
    ws.row_dimensions[rr].height = 20
r = blank(ws, r)
r = section(ws, r, "Backend / MAR allocation (NextLayer DC, Vienna)", NC)
headers = ["Resource", "Range / value", "Notes"]
r = header_row(ws, r, headers, [30, 30, 60])
be = [
    ("CDC management network", "10.178.0.0/16", "Server-side CDC networks (NextLayer DC)"),
    ("On-board management", "10.179.0.0/16", "Vehicle-side mgmt (OBS/CCU/switches) — CCUs land here"),
    ("Backend capacity", "10.178.0.0/17 + 10.179.0.0/17", "Up to 254 CCUs total; 6 Home Agents × 14 CCUs (120 Mbit/s each)"),
    ("MAR servers", "172.16.195.136 – .139", "mar3be06–mar3be09 (VPN home agents)"),
    ("NTP source (backend)", "10.178.13.1", "NTP for all IoB components"),
    ("Train→server (Dosto 1-10)", "→ 172.16.195.136 (mar3be06)", "NC-ID 10.178.1-10.x"),
    ("Train→server (Dosto 11-21)", "→ 172.16.195.137 (mar3be07)", "NC-ID 10.178.11-21.x"),
    ("Train→server (Dosto 22-32)", "→ 172.16.195.138 (mar3be08)", "NC-ID 10.178.22-32.x"),
    ("Train→server (Dosto 33-42 + benches)", "→ 172.16.195.139 (mar3be09)", "NC-ID 10.178.33-42.x + 96-120.x"),
]
r = data_rows(ws, r, be, mono_cols=(2,))
for rr in range(r-len(be), r):
    ws.row_dimensions[rr].height = 22
r = blank(ws, r)
r = note(ws, r, "CCUs have no direct Internet. All traffic is tunnelled via MAR5 VPN to the assigned Home Agent, filtered, and only then (if permitted) routed to the Internet. Servers are split across Home Agents so each hosts two consist types — a single server failure never strands a whole fleet type.", NC, 36)

# ----------------------------------------------------------------------------
# save
# ----------------------------------------------------------------------------
import sys as _sys
out = "reports/customer/DOSTO_NEU_Network_Communications_Matrix_v1.1.xlsx"
try:
    wb.save(out)
except PermissionError:
    out = "reports/customer/DOSTO_NEU_Network_Communications_Matrix_v1.1_new.xlsx"
    wb.save(out)
print("WROTE", out)
print("Tabs:", wb.sheetnames)
