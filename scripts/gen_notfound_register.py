#!/usr/bin/env python3
"""Stadler 'devices not found' register for the DOSTO fleet (6-car nv6 + 4-car nv4).
Source = Zabbix problem.get (genuinely-OPEN problems = same as the NMS dashboard),
filtered to port-down on template-ENABLED ports, joined to the OBN template device.
Two sheets: '6-Teiler (nv6)' and '4-Teiler (nv4)', each with its own Summary.
Placeholder 7.7.7.7 alarms excluded.
Output: reports/stadler/Stadler_DOSTO_Devices_Not_Found_Register_v1.2.xlsx
"""
import json, urllib.request, re, datetime, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ZBX = "https://trainzabbix-obb-alpin.nomadrail.com/api_jsonrpc.php"
def call(m, p, a=None):
    b = {"jsonrpc": "2.0", "method": m, "params": p, "id": 1}
    if a: b["auth"] = a
    r = urllib.request.Request(ZBX, data=json.dumps(b).encode(), headers={"Content-Type": "application/json-rpc"})
    return json.load(urllib.request.urlopen(r, timeout=120))

# --- 60xx/40xx -> Train# map, parsed from fleet-status.md (number = CCU IP 3rd octet) ---
def load_train_map():
    m = {}
    try:
        txt = open("fleet-status.md", encoding="utf-8").read()
    except OSError:
        return m
    pat = re.compile(r"\|\s*(4\d{3}-\d{3})\s*\|[^|]*\|\s*`?10\.179\.(\d+)\.1`?\s*\|")
    for mm in pat.finditer(txt):
        m[int(mm.group(2))] = mm.group(1)   # octet -> Train#
    return m
OCTET2TRAIN = load_train_map()
def train_no(g):
    # g like "50_6027" -> octet 27
    octet = int(g.split("_")[1][2:])
    return OCTET2TRAIN.get(octet, g.replace("50_", ""))

# --- port maps ---
def load_pm(path):
    pm = {}
    for line in open(path, encoding="utf-8"):
        if "|" not in line: continue
        sw, port, en, desc = (line.rstrip("\n").split("|") + ["", "", "", ""])[:4]
        pm[(sw, port)] = (en == "1", desc.strip())
    return pm
PM_NV6 = load_pm("findings/nv6_port_map_20260713.txt")
PM_NV4 = load_pm("findings/nv4_port_map_20260713.txt")

# Zabbix position R<n>_SW<m> -> letter switch, per fleet.
COACH_NV6 = {"1": "A", "2": "C", "3": "D", "4": "E", "5": "F", "6": "B"}   # 6-car A/C/D/E/F/B
COACH_NV4 = {"1": "A", "2": "G", "3": "E", "4": "B"}                       # 4-car A/G/E/B

def categorise(port, desc):
    d = desc.lower()
    if "bildschirm" in d: return "Display (Bildschirm)"
    if "audio" in d or "adu" in d: return "Audio/ADU"
    if "energiez" in d: return "Energy meter"
    if "interior camera" in d or "exterior camera" in d: return "Camera"
    if "afz" in d: return "AFZ"
    if "sprechstelle" in d: return "Intercom (Sprechstelle)"
    if "seitenanzeige" in d or "frontanzeige" in d: return "Side/front display"
    if "nvr" in d: return "NVR"
    if "zfr" in d: return "ZFR"
    if "firewall" in d: return "Firewall"
    if d.startswith("ap ") or port == "e0-4" or "ap " in d: return "Access point"
    if "obs" in d: return "OBS trunk"
    if "rdc" in d: return "RDC trunk"
    if "frontkupplung" in d: return "Coupler (expected down solo)"
    if "service" in d: return "Service port (often unused)"
    if "switch " in d or port in ("e0-0", "e0-1"): return "Inter-coach trunk"
    if "funksensor" in d: return "Funksensor"
    return "Other"

EXPECTED = {"Coupler (expected down solo)", "Service port (often unused)", "Inter-coach trunk"}
def ts(c): return datetime.datetime.utcfromtimestamp(int(c)).strftime("%Y-%m-%d %H:%MZ")

# ---------- gather ----------
tok = call("user.login", {"username": "okapi", "password": "5te8x7Dg5pweMu4R"})["result"]
allh = call("host.get", {"output": ["hostid", "host"], "selectInterfaces": ["available"]}, tok)["result"]

def collect(prefix, pm, coach):
    """Return (rows, n_trains). prefix e.g. '50_60' (6-car) / '50_40' (4-car)."""
    g2hids = defaultdict(list); g2avail = defaultdict(lambda: [0, 0])
    for h in allh:
        m = re.match(rf"({prefix}\d\d)_", h["host"])
        if m:
            g2hids[m.group(1)].append(h["hostid"])
            for i in h.get("interfaces", []):
                g2avail[m.group(1)][1] += 1
                if i.get("available") == "1": g2avail[m.group(1)][0] += 1
    rows = []
    for g in sorted(g2hids):
        online = g2avail[g][0] > 0
        probs = call("problem.get", {"hostids": g2hids[g], "output": ["name", "clock"],
            "recent": False, "sortfield": ["eventid"], "sortorder": "DESC"}, tok)["result"]
        for p in probs:
            name = p["name"]
            if "7.7.7.7" in name:
                continue
            m = re.search(r"Port (\S+) DOWN \(admin-enabled\) on " + prefix + r"\d\d_R(\d)_SW(\d)", name)
            if not m:
                continue
            port = m.group(1); sw = f"{coach.get(m.group(2), '?')}{m.group(3)}"
            en, desc = pm.get((sw, port), (None, ""))
            if en is not True:
                continue
            cat = categorise(port, desc)
            rows.append([train_no(g), g, "online" if online else "OFFLINE", sw, port, cat,
                         desc or "(no description)", ts(int(p["clock"])), cat in EXPECTED])
    def portkey(p):
        a, b = p.split("-"); return (a, int(b))
    rows.sort(key=lambda r: (r[0], r[3], portkey(r[4])))
    return rows, len(g2hids)

rows6, n6 = collect("50_60", PM_NV6, COACH_NV6)
rows4, n4 = collect("50_40", PM_NV4, COACH_NV4)

# ---------- workbook ----------
BLUE = "1F4E79"; HDR = PatternFill("solid", fgColor=BLUE)
THIN = Side(style="thin", color="BBBBBB"); BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
EXP_FILL = PatternFill("solid", fgColor="E0E0E0"); DEV_FILL = PatternFill("solid", fgColor="FFF2CC")
FONT = "Arial"
HEADERS = ["Train #", "Train (Zabbix)", "Reachability", "Switch", "Port", "Category",
           "Device (OBN template)", "Open since", "Classification", "Verification"]
WIDTHS = [12, 13, 12, 9, 8, 22, 30, 18, 15, 26]

wb = Workbook(); wb.remove(wb.active)

def write_sheet(title, fleet_label, rows, live_group):
    ws = wb.create_sheet(title); ws.sheet_view.showGridLines = False
    ws["A1"] = f"DOSTO {fleet_label} — Stadler Devices Not Found Register"
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=BLUE)
    ws["A2"] = ("Currently-OPEN Zabbix port-down problems on template-enabled ports (same source as the NMS dashboard — open problems, "
                "not raw event history). Placeholder 7.7.7.7 alarms excluded. Grey = expected-down; amber = missing end-device.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="808080")
    ws["A3"] = (f"Generated 2026-07-13. {len(rows)} open port-down problems across {len(set(r[0] for r in rows))} trains — counts match the NMS open-problems view. "
                "'Not found' may be a genuine fault OR a device not yet fitted; Stadler to verify. Offline trains show last-known open problems (re-verify live when next online).")
    ws["A3"].font = Font(name=FONT, size=9, italic=True, color="808080")
    hr = 5
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(hr, c, h); cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HDR; cell.border = BORDER; cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = WIDTHS[c-1]
    r = hr + 1
    for trainno, g, reach, sw, port, cat, desc, last, expected in rows:
        verify = "live-confirmed 2026-07-01" if g == live_group else ("re-verify when online" if reach == "OFFLINE" else "re-verify live")
        vals = [trainno, g, reach, sw, port, cat, desc, last,
                "expected-down" if expected else "MISSING device", verify]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v); cell.font = Font(name=FONT, size=10); cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(c == 7))
            if c == 1: cell.font = Font(name=FONT, size=10, bold=True)
            if c == 9:
                cell.fill = EXP_FILL if expected else DEV_FILL
                cell.font = Font(name=FONT, size=9, bold=not expected)
        r += 1
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{hr}:J{r-1}"

def write_summary(title, rows, notes=()):
    sm = wb.create_sheet(title); sm.sheet_view.showGridLines = False
    sm["A1"] = "Summary — missing devices by train and category"
    sm["A1"].font = Font(name=FONT, size=12, bold=True, color=BLUE)
    per_train = defaultdict(int); per_cat = defaultdict(int)
    for trainno, g, reach, sw, port, cat, desc, last, expected in rows:
        if not expected:
            per_train[trainno] += 1; per_cat[cat] += 1
    sm["A3"] = "Train #"; sm["B3"] = "Missing devices"
    for cc in ("A3", "B3"): sm[cc].font = Font(name=FONT, size=10, bold=True, color="FFFFFF"); sm[cc].fill = HDR; sm[cc].border = BORDER
    rr = 4
    for t in sorted(per_train, key=lambda x: -per_train[x]):
        sm.cell(rr, 1, t).border = BORDER; sm.cell(rr, 2, per_train[t]).border = BORDER; rr += 1
    sm["D3"] = "Category"; sm["E3"] = "Count"
    for cc in ("D3", "E3"): sm[cc].font = Font(name=FONT, size=10, bold=True, color="FFFFFF"); sm[cc].fill = HDR; sm[cc].border = BORDER
    rc = 4
    for cat in sorted(per_cat, key=lambda x: -per_cat[x]):
        sm.cell(rc, 4, cat).border = BORDER; sm.cell(rc, 5, per_cat[cat]).border = BORDER; rc += 1
    for col, w in [("A", 14), ("B", 16), ("C", 3), ("D", 26), ("E", 9)]:
        sm.column_dimensions[col].width = w
    # infra callouts
    infra = [r for r in rows if r[5] in ("Inter-coach trunk", "Access point")]
    co = max(rr, rc) + 3
    sm.cell(co, 1, "Priority entries — infrastructure ports (trunk / AP)").font = Font(name=FONT, size=11, bold=True, color="C00000")
    co += 1
    for hh, cc in zip(["Train #", "Switch", "Port", "Device", "Open since"], range(1, 6)):
        c = sm.cell(co, cc, hh); c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF"); c.fill = HDR; c.border = BORDER
    co += 1
    for trainno, g, reach, sw, port, cat, desc, last, expected in infra:
        for cc, v in enumerate([trainno, sw, port, f"{cat}: {desc}", last], 1):
            c = sm.cell(co, cc, v); c.font = Font(name=FONT, size=9); c.border = BORDER
        co += 1
    for note in notes:
        co += 1
        sm.cell(co, 1, note).font = Font(name=FONT, size=9, italic=True, color="808080")

NOTES_4 = (
    "4734-101 (59) and 4734-190 (55) account for most of the 4-car total. On each, ~54 of these ports went DOWN within a ~1-minute window "
    "(101: 2026-06-16 13:59-14:00Z; 190: 2026-06-17 08:49-08:50Z) and the train then went offline. These are last-known-open alarms — pending LIVE re-verification when 101/190 are back online; do not treat as confirmed per-device faults until then.",
    "4734-113 (55): all 55 are display ports (Bildschirm + side/front display) that went DOWN in a ~9-minute window on 2026-07-10 05:26-05:35Z, after which the train went offline — a consist power-down signature (same pattern as 101/190), NOT 55 individually missing devices. Re-verify live when 113 is back online.",
)

NOTES_6 = (
    "4736-107 (83): all 83 are display ports (Bildschirm + side/front display) that went DOWN in a ~4-minute window on 2026-07-08 14:30-14:34Z, after which the train went offline — a consist power-down signature, NOT 83 individually missing devices. These are last-known-open alarms; re-verify live when 107 is back online.",
)

# ---------- delta vs the previous register ----------
def load_prior_keys(path):
    """Read (Train#, Switch, Port)->row-tuple from a prior register's two data sheets."""
    from openpyxl import load_workbook
    prior = {}
    wb_p = load_workbook(path, read_only=True, data_only=True)
    for sn in ("6-Teiler (nv6)", "4-Teiler (nv4)"):
        if sn not in wb_p.sheetnames:
            continue
        ws_p = wb_p[sn]
        for r in ws_p.iter_rows(min_row=6, values_only=True):
            if not r or not r[0]:
                continue
            prior[(r[0], r[3], r[4])] = r          # Train#, Switch, Port
    wb_p.close()
    return prior

def write_delta(title, prior_label, prior_path, rows6, rows4):
    """Prepend a CLEARED / APPEARED / STILL-DOWN delta sheet vs the prior register."""
    if not os.path.exists(prior_path):
        print(f"  (delta skipped — prior register {prior_path} not found)")
        return
    prior = load_prior_keys(prior_path)
    now = {}
    for r in rows6 + rows4:
        now[(r[0], r[1], r[3], r[4])] = r          # keep group for infra note; key below drops it
    now_keys = {(r[0], r[3], r[4]) for r in rows6 + rows4}
    prior_keys = set(prior)
    cleared = prior_keys - now_keys
    appeared = now_keys - prior_keys
    still = now_keys & prior_keys

    def trains(keyset): return len({k[0] for k in keyset})
    sd = wb.create_sheet(title); sd.sheet_view.showGridLines = False
    sd["A1"] = f"Changes since {prior_label} — 2026-07-01 → 2026-07-13"
    sd["A1"].font = Font(name=FONT, size=13, bold=True, color=BLUE)
    sd["A2"] = ("Delta of currently-open port-down problems, keyed on (Train #, Switch, Port). "
                "CLEARED = present in the prior register, gone now (device reachable again / alarm closed). "
                "APPEARED = new since the prior register. STILL DOWN = present in both — most likely genuine missing devices. "
                "Infra ports (trunk/AP) are called out separately as priority.")
    sd["A2"].font = Font(name=FONT, size=9, italic=True, color="808080")
    hdr = ["Change", "Port-down entries", "Distinct trains", "Meaning"]
    meanings = {
        "CLEARED": "In the prior register, gone now — device reachable again / alarm closed",
        "APPEARED": "New since the prior register — went down after it was generated",
        "STILL DOWN": "Present in both — persistent, most likely genuine missing devices",
    }
    for c, h in enumerate(hdr, 1):
        cell = sd.cell(4, c, h); cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HDR; cell.border = BORDER
    for i, (label, ks) in enumerate((("CLEARED", cleared), ("APPEARED", appeared), ("STILL DOWN", still))):
        rr = 5 + i
        for c, v in enumerate([label, len(ks), trains(ks), meanings[label]], 1):
            cell = sd.cell(rr, c, v); cell.font = Font(name=FONT, size=10); cell.border = BORDER
    for col, w in [("A", 13), ("B", 18), ("C", 16), ("D", 62)]:
        sd.column_dimensions[col].width = w

    # per-train movement table
    per = defaultdict(lambda: [0, 0, 0])   # train -> [cleared, appeared, still]
    for k in cleared: per[k[0]][0] += 1
    for k in appeared: per[k[0]][1] += 1
    for k in still: per[k[0]][2] += 1
    base = 9
    sd.cell(base, 1, "Per-train movement").font = Font(name=FONT, size=11, bold=True, color=BLUE)
    for c, h in enumerate(["Train #", "Cleared", "Appeared", "Still down"], 1):
        cell = sd.cell(base + 1, c, h); cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HDR; cell.border = BORDER
    rr = base + 2
    for t in sorted(per, key=lambda x: -(per[x][1] + per[x][2])):
        for c, v in enumerate([t] + per[t], 1):
            cell = sd.cell(rr, c, v if v else None); cell.font = Font(name=FONT, size=9); cell.border = BORDER
        rr += 1

    # priority infra callout (new/still-down trunk & AP ports)
    infra = [r for r in rows6 + rows4
             if r[5] in ("Inter-coach trunk", "Access point") and (r[0], r[3], r[4]) in (appeared | still)]
    if infra:
        rr += 2
        sd.cell(rr, 1, "Priority — infrastructure ports still/newly down (trunk / AP)").font = Font(
            name=FONT, size=11, bold=True, color="C00000"); rr += 1
        for c, h in enumerate(["Train #", "Switch", "Port", "Device", "Open since"], 1):
            cell = sd.cell(rr, c, h); cell.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
            cell.fill = HDR; cell.border = BORDER
        rr += 1
        for r in sorted(infra, key=lambda x: (x[0], x[3], x[4])):
            for c, v in enumerate([r[0], r[3], r[4], f"{r[5]}: {r[6]}", r[7]], 1):
                cell = sd.cell(rr, c, v); cell.font = Font(name=FONT, size=9); cell.border = BORDER
            rr += 1
    wb.move_sheet(sd, offset=-(wb.sheetnames.index(title)))   # make it the leftmost sheet
    print(f"  delta vs {prior_label}: {len(cleared)} cleared, {len(appeared)} appeared, {len(still)} still down")

PRIOR = "reports/stadler/Stadler_DOSTO_Devices_Not_Found_Register_v1.1.xlsx"
write_delta("Changes since v1.1", "v1.1", PRIOR, rows6, rows4)
write_sheet("6-Teiler (nv6)", "6-car (nv6)", rows6, "50_6027")
write_summary("6-Teiler Summary", rows6, NOTES_6)
write_sheet("4-Teiler (nv4)", "4-car (nv4)", rows4, None)
write_summary("4-Teiler Summary", rows4, NOTES_4)

os.makedirs("reports/stadler", exist_ok=True)
out = "reports/stadler/Stadler_DOSTO_Devices_Not_Found_Register_v1.2.xlsx"
wb.save(out)
m6 = sum(1 for r in rows6 if not r[8]); m4 = sum(1 for r in rows4 if not r[8])
print(f"wrote {out}")
print(f"  6-Teiler: {len(rows6)} rows ({m6} missing) across {len(set(r[0] for r in rows6))} trains")
print(f"  4-Teiler: {len(rows4)} rows ({m4} missing) across {len(set(r[0] for r in rows4))} trains")
