#!/usr/bin/env python3
"""Append a 'Changes since v1.0' sheet to the v1.1 Devices Not Found register.
Diffs v1.0 (2026-06-20) vs v1.1 (2026-07-01) by (Train#, Switch, Port):
  CLEARED   = in v1.0, gone in v1.1  (device now reachable / alarm closed)
  APPEARED  = new in v1.1            (went down since 2026-06-20)
  PERSIST   = in both                (still down)
Adds one detail table (appeared+cleared per train) and a priority infra callout.
Writes the sheet back into the v1.1 workbook in place.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

V10 = "reports/stadler/Stadler_DOSTO_Devices_Not_Found_Register_v1.0.xlsx"
V11 = "reports/stadler/Stadler_DOSTO_Devices_Not_Found_Register_v1.1.xlsx"

BLUE = "1F4E79"; HDR = PatternFill("solid", fgColor=BLUE)
THIN = Side(style="thin", color="BBBBBB"); BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GREEN = PatternFill("solid", fgColor="C6EFCE"); AMBER = PatternFill("solid", fgColor="FFF2CC")
GREY = PatternFill("solid", fgColor="E0E0E0")
FONT = "Arial"

def load(path):
    wb = load_workbook(path, data_only=True); rows = []
    for sh in ["6-Teiler (nv6)", "4-Teiler (nv4)"]:
        for r in wb[sh].iter_rows(min_row=6, values_only=True):
            if r[0] is None:
                continue
            rows.append({"train": r[0], "reach": r[2], "sw": r[3], "port": r[4],
                         "cat": r[5], "dev": r[6], "since": r[7], "cls": r[8]})
    return rows

old = load(V10); new = load(V11)
def key(r): return (r["train"], r["sw"], r["port"])
oldk = {key(r): r for r in old}; newk = {key(r): r for r in new}
cleared = [oldk[k] for k in oldk if k not in newk]
appeared = [newk[k] for k in newk if k not in oldk]
persist = [newk[k] for k in newk if k in oldk]

wb = load_workbook(V11)
if "Changes since v1.0" in wb.sheetnames:
    del wb["Changes since v1.0"]
ws = wb.create_sheet("Changes since v1.0", 0)   # front of workbook
ws.sheet_view.showGridLines = False

ws["A1"] = "Changes since v1.0 — 2026-06-20 → 2026-07-01"
ws["A1"].font = Font(name=FONT, size=14, bold=True, color=BLUE)
ws["A2"] = ("Delta of currently-open port-down problems, keyed on (Train #, Switch, Port). "
            "Most CLEARED entries are trains that were offline / last-known-open on 2026-06-20 and have since come back online healthy. "
            "APPEARED = went down after 2026-06-20. Infra ports (trunk/AP) are called out separately as priority.")
ws["A2"].font = Font(name=FONT, size=9, italic=True, color="808080")

def head(row, cols, widths):
    for c, (h, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row, c, h); cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HDR; cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w

# --- top-line counts ---
r = 4
ws.cell(r, 1, "Category").font = Font(name=FONT, size=11, bold=True, color=BLUE); r += 1
head(r, ["Change", "Port-down entries", "Distinct trains", "Meaning"], [16, 18, 16, 60]); r += 1
for label, data, fill, meaning in [
    ("CLEARED", cleared, GREEN, "In v1.0, gone now — device reachable again / alarm closed (mostly recovered offline trains)"),
    ("APPEARED", appeared, AMBER, "New since 2026-06-20 — went down after the last register"),
    ("STILL DOWN", persist, GREY, "Present in both — persistent, most likely genuine missing devices"),
]:
    trains = len(set(x["train"] for x in data))
    for c, v in enumerate([label, len(data), trains, meaning], 1):
        cell = ws.cell(r, c, v); cell.font = Font(name=FONT, size=10, bold=(c == 1)); cell.border = BORDER
        if c == 1: cell.fill = fill
        cell.alignment = Alignment(wrap_text=(c == 4), vertical="center")
    r += 1

# --- per-train appeared/cleared ---
r += 2
ws.cell(r, 1, "Per-train movement").font = Font(name=FONT, size=11, bold=True, color=BLUE); r += 1
head(r, ["Train #", "Cleared", "Appeared", "Still down", "Note"], [12, 10, 11, 12, 50]); r += 1
pc = defaultdict(int); pa = defaultdict(int); pp = defaultdict(int)
for x in cleared: pc[x["train"]] += 1
for x in appeared: pa[x["train"]] += 1
for x in persist: pp[x["train"]] += 1
alltrains = set(pc) | set(pa) | set(pp)
NOTE = {
    "4734-190": "Recovered — 54 of 55 cleared (was the v1.0 last-known-open cluster)",
    "4734-101": "NOT recovered — 59 still down; promote from last-known-open to verify-for-real",
    "4736-112": "Recovered — 82 cleared",
    "4736-113": "Mostly recovered",
    "4706-102": "Whole train appeared — 17 devices, likely first online / not yet fitted",
    "4736-119": "New inter-coach trunk down both ends (B2 e0-0 / B3 e0-1) — priority",
}
for t in sorted(alltrains, key=lambda x: -(pa[x] + pp[x])):
    for c, v in enumerate([t, pc[t] or "", pa[t] or "", pp[t] or "", NOTE.get(t, "")], 1):
        cell = ws.cell(r, c, v); cell.font = Font(name=FONT, size=9, bold=(c == 1)); cell.border = BORDER
        cell.alignment = Alignment(wrap_text=(c == 5), vertical="center")
    r += 1

# --- priority infra among appeared ---
r += 2
ws.cell(r, 1, "Priority — new infrastructure faults (trunk / AP) since 2026-06-20").font = \
    Font(name=FONT, size=11, bold=True, color="C00000"); r += 1
head(r, ["Train #", "Switch", "Port", "Category", "Device", "Open since"], [12, 9, 8, 20, 22, 18]); r += 1
infra = sorted([x for x in appeared if x["cat"] in ("Inter-coach trunk", "Access point")],
               key=lambda x: (x["cat"] != "Inter-coach trunk", x["train"]))
for x in infra:
    for c, v in enumerate([x["train"], x["sw"], x["port"], x["cat"], x["dev"], x["since"]], 1):
        cell = ws.cell(r, c, v); cell.font = Font(name=FONT, size=9); cell.border = BORDER
    r += 1

ws.freeze_panes = "A3"
wb.save(V11)
print(f"added 'Changes since v1.0' sheet to {V11}")
print(f"  cleared={len(cleared)} appeared={len(appeared)} persist={len(persist)}")
print(f"  priority infra appeared={len(infra)}")
