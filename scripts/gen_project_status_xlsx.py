#!/usr/bin/env python3
"""Generate PROJECT-STATUS.xlsx — flat single-sheet master tracker for the DOSTO project.
Source: PROJECT-STATUS.md (8 workstreams) + 2026-06-20 session updates (4736-114, cable reg #11)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = "1F4E79"
HDR_FILL = PatternFill("solid", fgColor=BLUE)
SUM_FILL = PatternFill("solid", fgColor="D5E8F0")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT = "Arial"

# Status palette
STATUS_FILL = {
    "Done":        PatternFill("solid", fgColor="C6EFCE"),
    "In progress": PatternFill("solid", fgColor="FFEB9C"),
    "Todo":        PatternFill("solid", fgColor="FFFFFF"),
    "Blocked":     PatternFill("solid", fgColor="FFC7CE"),
    "Deferred":    PatternFill("solid", fgColor="E0E0E0"),
}

# columns: Workstream, Item, Owner, Status, Detail, Last updated
ROWS = [
    # --- 1. Commissioning ---
    ("Commissioning", "v8 rollout — fleet (per-train detail in fleet-status.md)", "Nomad", "In progress",
     "4736: 11 done, 4 paused, 1 error (118 PoE), 7 unknown. 4734: 8 done, 6 paused, 1 blocked. 4706/4705: mostly unknown.", "2026-06-20"),
    ("Commissioning", "4736-114 / Fzg 142 — misimage corrected", "Nomad", "In progress",
     "RESOLVED misimage (was imaged as 4706-103/T17). Now box1-t27 @ 10.179.27.1, 18/18 sw v8-142, 24/24 AP, vlan7 .199.2/17 carrying traffic, Zabbix 6027 healthy. Remaining: AP fw push 0/24 -> 6.11.2-0.", "2026-06-20"),
    ("Commissioning", "AP firmware partition-swap wall", "Nomad/R&D", "Todo",
     "Many staged-not-activated APs across fleet; obn update f stages fw but reboot doesn't activate. Bug11 now reports honestly; per-AP retry + LuCI bypass. R&D candidate.", "2026-06-20"),
    ("Commissioning", "4705 platform unsupported by commission skill", "Nomad", "Blocked",
     "dosto-commission-train only enumerates 4-car/6-car; fv5-* (4705) needs skill update before commissioning.", "2026-05-22"),
    ("Commissioning", "OBN persisted-baseline drift (top-up bug10/11)", "Nomad", "Todo",
     "Rows claiming 8/8 or 9/9 are silently <=9/11 or <=10/11. 14 pre-bug10 + 6 pre-bug11 trains need top-up. Low urgency.", "2026-06-09"),

    # --- 2. SDD ---
    ("SDD", "SDD-002 T1 — QoS/DSCP per-VLAN write-back", "Nomad", "Todo",
     "Produces v2.4. The one item closable with no external input.", "2026-06-19"),
    ("SDD", "SDD-003 — 15 mechanical comments", "Nomad", "Done",
     "All 15 done in v3-tracked.", "2026-06-19"),
    ("SDD", "SDD-003 — 10 comments blocked on inputs (T2-T11)", "Nomad/OEBB", "Blocked",
     "Dimensioning sheet (biggest unblocker), NAT ratio, Wi-Fi channel plan, VLAN47, Project-ID confirm, diagrams, portal screenshots/spec/naming, NTP source.", "2026-06-19"),
    ("SDD", "SDD-003 #154 — commercial dispute", "Other", "Deferred",
     "Handled separately by email.", "2026-06-19"),

    # --- 3. Cabling / hardware (Stadler) ---
    ("Cabling/HW", "Reg #1 — 4734-101 E2<->B1 wrong neighbour", "Stadler", "Blocked", "OPEN — re-patch E<->B trunk.", "2026-05-22"),
    ("Cabling/HW", "Reg #4 — 4736-109 B3 e0-4 AP not connected", "Stadler", "Blocked", "OPEN.", "2026-06-09"),
    ("Cabling/HW", "Reg #5 — 4736-104 D3 e1-2 physical-layer", "Stadler", "Blocked", "OPEN — data pairs dead, cable replace.", "2026-05-22"),
    ("Cabling/HW", "Reg #7 — 4736-108 A2<->A3 e0-1 physical-layer", "Stadler", "Blocked", "OPEN — RX errors under load; SFP/cable. Caused storm.", "2026-05-29"),
    ("Cabling/HW", "Reg #8 — 4736-115 B3 e0-4 Coach6 AP not connected", "Stadler", "Blocked", "OPEN.", "2026-06-09"),
    ("Cabling/HW", "Reg #9 — 4736-118 E1 PoE PSE fault", "Stadler", "In progress", "MONITORING — was OPEN (dead PoE PSE). 2026-06-20 live check: E1 PoE HEALTHY (33W, same chassis, SNMP up); fault not currently present but was intermittent. Confirm stable across power-cycles before closing.", "2026-06-20"),
    ("Cabling/HW", "Reg #11 — 4736-114 14x FIS display/energy-meter ports", "OEBB/Stadler", "Blocked",
     "OPEN — 14 e2-* display/meter ports link-down. CCU-side EXHAUSTED (port details, 40 up siblings, bounce x14, logs clean). Verification requested via OEBB report 2026-06-20. May be fit-out state not faults. Zabbix 6027 alarms match exactly.", "2026-06-20"),
    ("Zabbix/NMS", "Coverage gap — 4734-101 (6004) absent from Zabbix", "Nomad", "Todo",
     "Cable reg #1 (4734-101 E2<->B1) cannot be alarmed — train not provisioned in Zabbix (no 50_6004 group). 'No alarm' != 'no fault'. Provision 6004.", "2026-06-20"),
    ("Zabbix/NMS", "Coverage gap — no RX-error-rate trigger (trunk degradation invisible)", "Nomad", "Todo",
     "Cable reg #7 (4736-108 A2<->A3 error trunk, link UP) is invisible in Zabbix — only port-down/SNMP-unreachable triggers exist, no ifInErrors-rate. Add error-rate item/trigger to switch template.", "2026-06-20"),

    # --- 4. OBN bugs ---
    ("OBN bugs", "11-bug suite characterised + dosto-obn-patches", "Nomad", "Done", "Bugs 1-10 code + bug11 AP-fw verify; applied/verified by skill.", "2026-06-09"),
    ("OBN bugs", "nd-obn 2.2.23 native fixes understood", "Nomad", "Done", "2.2.23 ships bugs 1-10 native; only bug11 hand-patch; never run fix_obn.py on it.", "2026-06-10"),
    ("OBN bugs", "TRIAG-8585 — upstream R&D ticket", "R&D", "In progress", "Tracks 11 code + 1 infra bug for v8 hand-patches.", "2026-06-09"),
    ("OBN bugs", "TFTP conntrack helper — Puppet fix", "R&D", "Blocked", "Runtime modprobe+iptables works but wiped on reboot; needs Puppet in 60-allow-management.", "2026-06-09"),
    ("OBN bugs", "dosto-ap-firmware-update skill precheck bugs", "Nomad", "Todo", "snmpget too strict; journalctl ISO date.", "2026-06-09"),
    ("OBN bugs", "MAR5 hostname equivalents from R&D", "R&D", "Todo", "Awaiting R&D.", "2026-06-09"),

    # --- 5. Zabbix / NMS ---
    ("Zabbix/NMS", "SNMP cred model resolved", "Nomad", "Done", "SW user=snmpadmin, AP user=admin, both NomadStayOut!.", "2026-06-08"),
    ("Zabbix/NMS", "NV6 template AP-user fix (v9)", "Nomad", "Done", "securityName snmpadmin->admin. NV4/FV5/bench unchecked.", "2026-06-08"),
    ("Zabbix/NMS", "Switch template wrong OIDs corrected", "Nomad", "Done", "port-status + firmware OIDs fixed.", "2026-06-08"),
    ("Zabbix/NMS", "NMS CDC page empty (Alpin) — stale OVH creds", "Nomad/sysops", "Blocked", "PARKED until Mon 2026-06-22 on sysops creds; then fix both projects.", "2026-06-20"),
    ("Zabbix/NMS", "Proxy CacheSize/LLD bloat", "Nomad", "In progress", "ifDescr over-discovery OOM; {#IFDESCR} filter added, CacheSize=128M stopgap.", "2026-06-09"),
    ("Zabbix/NMS", "Stale-template-IP reconciliation", "Nomad", "In progress", "Reconcile SNMP-dead host IPs from OBN discovery.", "2026-06-09"),

    # --- 6. 6040 / GPS ---
    ("6040/GPS", "6040 GPS blank in NMS — rtl_project_id 51 not 50", "Nomad", "Blocked", "Publish namespace stuck p51; OEBB NMS monitors p50. Fix in facter/telemetry/bridge.", "2026-06-15"),
    ("6040/GPS", "6040 link-info gap", "Nomad", "Done", "Was post-reboot modem reidentify transient; self-healed.", "2026-06-15"),
    ("6040/GPS", "GPS MQTT bridge 'mismatch' — RETRACTED", "Nomad", "Done", "Bridge is correct; gap is NMS-side. Do not touch bridge.", "2026-06-15"),

    # --- 7. Open investigations ---
    ("Investigations", "Fzg 19/20 B&E WAP swap", "Nomad", "In progress", "Customer-reported WE1/WE2 swap; ruled out OBN/cabling/config; suspect SSID config or coach orientation.", "2026-06-15"),
    ("Investigations", "Fzg 1 partial health check", "Nomad", "In progress", "Partial L2 sweep box1-t4; CCU no sshpass, RO rootfs; 2 APs missing; train offline mid-sweep.", "2026-06-15"),
    ("Investigations", "Fzg136 obn user-count slow-SNMP", "Nomad", "Done", "Slow but completes every cycle; not stuck, no action.", "2026-06-15"),

    # --- 8. Tooling / infra ---
    ("Tooling", "Auto-scanner", "Nomad", "In progress", "Tier-1 reachability stub built+tested; Tier-2 + bootstrap deferred.", "2026-06-09"),
    ("Tooling", "Orchestration stack", "Nomad", "Done", "Inline /dosto-orchestrate + worker + 19-stage commission-train + 4 contracts. Feature-complete.", "2026-06-09"),
    ("Tooling", "Jira integration decision", "Nomad/team", "Blocked", "Surveyed EMEAE; 3 options. Blocked on team coordination (Kemeter/Behnam/Zejnelovic).", "2026-05-10"),
]

HEADERS = ["Workstream", "Item", "Owner", "Status", "Detail", "Last updated"]
WIDTHS = [16, 48, 14, 13, 70, 14]

wb = Workbook()

# ---------- Status sheet ----------
ws = wb.active
ws.title = "Status"
ws.sheet_view.showGridLines = False

# Title
ws["A1"] = "DOSTO Project — Master Status Tracker"
ws["A1"].font = Font(name=FONT, size=15, bold=True, color=BLUE)
ws["A2"] = "Single source of truth across all workstreams. Filter/sort by Workstream or Status. Update the row + Last updated when state changes."
ws["A2"].font = Font(name=FONT, size=9, italic=True, color="808080")
ws["A3"] = "Generated 2026-06-20 (replaces PROJECT-STATUS.md). Per-train detail still lives in fleet-status.md; cabling detail in cable-issues-register.md."
ws["A3"].font = Font(name=FONT, size=9, italic=True, color="808080")

hdr_row = 5
for c, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=hdr_row, column=c, value=h)
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(c)].width = WIDTHS[c-1]

r = hdr_row + 1
for ws_name, item, owner, status, detail, updated in ROWS:
    vals = [ws_name, item, owner, status, detail, updated]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name=FONT, size=10)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=(c == 5))
        cell.border = BORDER
        if c == 4:  # status colour
            cell.fill = STATUS_FILL.get(status, STATUS_FILL["Todo"])
            cell.font = Font(name=FONT, size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

last_data_row = r - 1
ws.freeze_panes = "A6"
ws.auto_filter.ref = f"A{hdr_row}:F{last_data_row}"

# ---------- Summary sheet ----------
sm = wb.create_sheet("Summary")
sm.sheet_view.showGridLines = False
sm["A1"] = "Status summary"
sm["A1"].font = Font(name=FONT, size=13, bold=True, color=BLUE)

# counts by status (formulas against Status sheet)
statuses = ["Done", "In progress", "Todo", "Blocked", "Deferred"]
sm["A3"], sm["B3"] = "Status", "Count"
for c, t in [("A3", "Status"), ("B3", "Count")]:
    sm[c].font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    sm[c].fill = HDR_FILL
    sm[c].border = BORDER
status_counts = {s: sum(1 for row in ROWS if row[3] == s) for s in statuses}
for i, s in enumerate(statuses):
    rr = 4 + i
    sm.cell(row=rr, column=1, value=s).font = Font(name=FONT, size=10)
    sm.cell(row=rr, column=1).fill = STATUS_FILL[s]
    sm.cell(row=rr, column=1).border = BORDER
    cc = sm.cell(row=rr, column=2, value=status_counts[s])
    cc.font = Font(name=FONT, size=10)
    cc.border = BORDER
total_row = 4 + len(statuses)
sm.cell(row=total_row, column=1, value="Total").font = Font(name=FONT, size=10, bold=True)
sm.cell(row=total_row, column=1).fill = SUM_FILL
sm.cell(row=total_row, column=1).border = BORDER
tc = sm.cell(row=total_row, column=2, value=len(ROWS))
tc.font = Font(name=FONT, size=10, bold=True)
tc.fill = SUM_FILL
tc.border = BORDER

# counts by workstream
sm["D3"], sm["E3"] = "Workstream", "Count"
for c in ("D3", "E3"):
    sm[c].font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    sm[c].fill = HDR_FILL
    sm[c].border = BORDER
workstreams = ["Commissioning", "SDD", "Cabling/HW", "OBN bugs", "Zabbix/NMS", "6040/GPS", "Investigations", "Tooling"]
ws_counts = {w: sum(1 for row in ROWS if row[0] == w) for w in workstreams}
for i, wsn in enumerate(workstreams):
    rr = 4 + i
    sm.cell(row=rr, column=4, value=wsn).font = Font(name=FONT, size=10)
    sm.cell(row=rr, column=4).border = BORDER
    cc = sm.cell(row=rr, column=5, value=ws_counts[wsn])
    cc.font = Font(name=FONT, size=10)
    cc.border = BORDER

for col, w in [("A", 14), ("B", 9), ("C", 3), ("D", 16), ("E", 9)]:
    sm.column_dimensions[col].width = w

out = "PROJECT-STATUS.xlsx"
wb.save(out)
print("wrote", out, "with", len(ROWS), "items")
