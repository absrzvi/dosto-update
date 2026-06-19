#!/usr/bin/env python3
"""
fill_zug_land.py — Populate the ZUG-LAND (train-to-ground) connection matrix in the
DOSTO NEU network-planning workbook.

Reads the frozen design-freeze workbook, fills the ZUG-LAND sheet's tbd rows with every
train<->land flow we have an authoritative source for, marks genuinely-open items as
"tbd (Stadler)" with the owner named, and writes a NEW versioned working copy
(v0.8 WORKING) so the frozen v7 is left untouched.

Sources (all from the same workbook + design-freeze SDD v2.2):
  - 'Erlaubte Verbindungen' (SDD §6.18.1 / Network tab)      -> Stadler wayside connections
  - 'NC-ID und Fzg-ID' sheet                                  -> MAR5 backend public IPs (mar3be06-09)
  - 'IP-Adresse LAND' sheet                                   -> AFZ & CCTV landside servers
  - 'IP-Adresse ZUG' sheet                                    -> T2G CCU/trainserver VIPs
  - SDD 'Backend Netzwerke' table                             -> backend NTP 10.178.13.1
  - SDD 'Zeit synchronisation'                                -> NTP / DNS flows
"""
import copy
import openpyxl
from openpyxl.styles import Alignment

SRC = "design freeze/network-planning-v7.xlsx"
OUT = "design freeze/network-planning-v8-WORKING.xlsx"

wb = openpyxl.load_workbook(SRC)   # keep formatting/themes
ws = wb["ZUG-LAND"]

# ---- capture the existing cell styles to mirror the sheet's house style ----
# Header style sample (row 2 = column headers, bold white on theme fill, bordered)
hdr_style_cell = ws["A2"]
# Placeholder/data style sample (row 3 = the old 'tbd' cells, bordered, plain)
data_style_cell = ws["A3"]

def style_like(target, src_cell, *, wrap=True, align_h="left", align_v="center"):
    target.font = copy.copy(src_cell.font)
    target.fill = copy.copy(src_cell.fill)
    target.border = copy.copy(src_cell.border)
    target.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)

# Columns:
# A Description | B Direction | C Endpoint A (Source) | D IP | E Port
# F Endpoint B (Destination) | G IP | H Port
# I Data-Link (VLAN) | J Network-Layer (proto) | K Transport-Layer (proto)
COLS = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]

# Each row: (Description, Direction, SrcEndpoint, SrcIP, SrcPort, DstEndpoint, DstIP, DstPort, VLAN, NetL3, TransL4)
# "Train" direction = originated on-board, towards land.
rows = []

# ---- Section marker rows (we colour them like a sub-header) ----
def section(label):
    rows.append(("__SECTION__", label))

# === A. Nomad IoB — CCU to MAR5 backend (the primary IoB train<->ground tunnel) ===
# NOTE: the CCU connects to its ASSIGNED MAR5 Home Agent. The HA is chosen by Dosto-Neu
# sequence number (SDD Table 8), NOT by fleet family, and one HA serves a mix of types.
# We therefore present the tunnel generically and list the 4 HAs as a reference block,
# rather than asserting a (source-inconsistent) per-family mapping.
section("Nomad IoB — CCU ↔ MAR5 Backend (NextLayer DC, Wien)  ·  source: SDD Table 8 + 'NC-ID und Fzg-ID'")
rows += [
    ("MAR5 VPN tunnel (CCU → assigned HA)", "Train → Land", "CCU (bond0)", "10.179.x.1 / 10.178.x.1", "ephemeral",
     "Assigned MAR5 Home Agent", "mar3be06–09 (see ref below)", "UDP/MAR5", "WAN (cellular)", "IPv4", "UDP (MAR5 tunnel, AES-256)"),
    ("· HA mar3be06  (Dosto Neu 1–10)", "reference", "—", "—", "—",
     "mar3be06.nextlayer", "int 172.16.195.136 / pub 213.208.157.7", "—", "—", "IPv4", "—"),
    ("· HA mar3be07  (Dosto Neu 11–21)", "reference", "—", "—", "—",
     "mar3be07.nextlayer", "int 172.16.195.137 / pub 213.208.157.9", "—", "—", "IPv4", "—"),
    ("· HA mar3be08  (Dosto Neu 22–32)", "reference", "—", "—", "—",
     "mar3be08.nextlayer", "int 172.16.195.138 / pub 213.208.157.77", "—", "—", "IPv4", "—"),
    ("· HA mar3be09  (Dosto Neu 33–42 + benches)", "reference", "—", "—", "—",
     "mar3be09.nextlayer", "int 172.16.195.139 / pub 213.208.157.79", "—", "—", "IPv4", "—"),
    ("Telemetry / MQTT (via tunnel)", "Train → Land", "CCU (telemetry)", "10.179.x.1", "ephemeral",
     "HiveMQ MQTT broker (backend)", "via MAR5 HA", "1883 / 8883", "100 → WAN", "IPv4 (in tunnel)", "TCP (MQTT/MQTTS)"),
    ("NMS / Telemetry stats, logs", "Train → Land", "CCU", "10.179.x.1", "ephemeral",
     "NMS app / Zabbix / RTL log collector", "CDC 10.178.0.0/16", "various", "100 → WAN", "IPv4 (in tunnel)", "TCP"),
    ("Portal / OBN image pull, updates", "Land → Train", "Rollout Mgr / CDC", "CDC 10.178.0.0/16", "443",
     "CCU (Docker portal)", "10.179.x.1", "443", "WAN → 100", "IPv4 (in tunnel)", "TCP (HTTPS)"),
    ("Engineering Pages access (staff)", "On-board", "iobstaff client", "10.205.{train}.x", "ephemeral",
     "CCU Engineering Pages", "engineering.page", "8005", "30", "IPv4", "TCP (HTTPS)"),
]

# === B. Stadler wayside connections (the only Stadler->land path, via VLAN 7) ===
section("Stadler wayside — routed via VLAN 7 → CCU (NAT) → Land  ·  source: SDD §6.18.1 'Erlaubte Verbindungen'")
rows += [
    ("Stadler RDS VPN tunnel", "Train → Land", "Stadler Router (via vlan7)", "172.19.x.x (vlan7)", "ephemeral",
     "Stadler RDS", "62.2.130.53", "83", "7 → WAN", "IPv4", "UDP"),
    ("Stadler RDS VPN tunnel (2nd port)", "Train → Land", "Stadler Router (via vlan7)", "172.19.x.x (vlan7)", "ephemeral",
     "Stadler RDS", "62.2.130.53", "84", "7 → WAN", "IPv4", "UDP"),
    ("Data transfer to ÖBB landside", "Train → Land", "Stadler Router (via vlan7)", "172.19.x.x (vlan7)", "ephemeral",
     "ÖBB landside (SFTP)", "195.69.192.153", "22", "7 → WAN", "IPv4", "TCP (SSH/SFTP)"),
    ("IPSec tunnel to video server", "Train → Land", "Stadler Router (via vlan7)", "172.19.x.x (vlan7)", "500",
     "Video server (Stadler)", "tbd (Stadler)", "500", "7 → WAN", "IPv4", "UDP (IKE/IPSec)"),
    ("AFZ → landside AFZ server", "Train → Land", "AFZ (via vlan7/SecGW)", "VLAN 8 (Stadler)", "ephemeral",
     "ftp-afz.landside.oebb-flotte", "195.69.194.146", "tbd (Stadler)", "8 → 7 → WAN", "IPv4", "tbd (FTP/SFTP)"),
    ("CCTV → landside CCTV server", "Train → Land", "Recorder (via vlan7/SecGW)", "VLAN 5 (Stadler)", "ephemeral",
     "cctv.landside.oebb-flotte", "10.10.x.y (Stadler landing-zone)", "tbd (Stadler)", "5 → 7 → WAN", "IPv4", "tbd (Stadler)"),
]

# === C. Time / name services across the boundary ===
section("Time & name services  ·  source: SDD 'Zeit synchronisation' + 'Backend Netzwerke'")
rows += [
    ("NTP — OBS reference (primary)", "Train → Land", "OBS (NTP master)", "VLAN 7 (Stadler)", "123",
     "Landside Internet NTP", "via vlan7 → CCU → WAN", "123", "7 → WAN", "IPv4", "UDP (NTP)"),
    ("NTP — backend source (IoB)", "Train → Land", "CCU / switches / APs", "10.179.x.x", "123",
     "Backend NTP server", "10.178.13.1", "123", "100 → WAN", "IPv4 (in tunnel)", "UDP (NTP)"),
    ("NTP — RCU via SNAT (3rd-party)", "On-board", "Stadler RCU NTP", "192.168.101.10 / .15", "123",
     "OBS (via Stadler FW SNAT)", "VLAN 7", "123", "7", "IPv4", "UDP (NTP, SNAT by FW)"),
    ("DNS — zug-domain resolution", "On-board / Land", "Train clients / CCU", "per VLAN", "ephemeral",
     "DNS (CCU + backend blacklisting)", "CCU + CDC DNS", "53", "various → WAN", "IPv4", "TCP/UDP (DNS)"),
    ("GPS / NMEA feed (staff/CCU)", "On-board", "CCU GPS source", "10.179.x.1", "ephemeral",
     "iobstaff / OBS (vlan7)", "10.205.x.x / vlan7", "2947", "30 / 7", "IPv4", "TCP (gpsd)"),
]

# === D. T2G (train-to-ground) reference addressing — Stadler/RDC model (per IP-Adresse sheets) ===
section("Train-to-Ground (T2G) reference addresses  ·  source: 'IP-Adresse ZUG' / 'IP-Adresse LAND' (Stadler/RDC model)")
rows += [
    ("CCU T2G VIP (router addr)", "reference", "ccu-t2g", "10.70.0.13", "n/a",
     "(T2G VIP for train↔land)", "—", "n/a", "T2G", "IPv4", "—"),
    ("CCTV recorder external access", "reference", "rec1b / rec2b", "10.52.0.10 / .11", "n/a",
     "(external CCTV access VIP)", "—", "n/a", "T2G", "IPv4", "—"),
    ("AFZ passenger counting", "reference", "AFZ", "10.38.0.10", "n/a",
     "ftp-afz.landside.oebb-flotte", "195.69.194.146", "n/a", "T2G", "IPv4", "—"),
    ("CCU backend servers (T2G DNS)", "reference", "mar3be1-4.nextlayer.oebb.com", "see MAR5 alloc", "n/a",
     "(backend home agents)", "213.208.157.7/.9/.77/.79", "n/a", "T2G", "IPv4", "—"),
]

# === E. Genuinely-open items (Stadler-owned, per INFO 'Offene Themen') ===
section("Open items — owner: Stadler (INFO 'Offene Themen', 2025-11-03)")
rows += [
    ("Zug-Land Endpunkte (full list)", "tbd", "tbd (Stadler)", "tbd (Stadler)", "tbd",
     "tbd (Stadler)", "tbd (Stadler)", "tbd", "tbd", "tbd", "tbd"),
    ("Multitraktion Stadler-VLAN routing", "tbd", "tbd (Stadler/Nomad VO)", "tbd", "tbd",
     "tbd", "tbd", "tbd", "5,15 (+VO)", "tbd", "tbd"),
    ("Reservierung / PIS wayside endpoints", "tbd", "tbd (Stadler)", "tbd (Stadler)", "tbd",
     "tbd (Stadler)", "tbd (Stadler)", "tbd", "6 / 3", "tbd", "tbd"),
]

# ---------------------------------------------------------------------------
# Clear the old 'tbd' placeholder rows (3-5) and write the new content.
# ---------------------------------------------------------------------------
# wipe rows 3..(existing max) in cols A..K
for rr in range(3, ws.max_row + 1):
    for col in COLS:
        ws[f"{col}{rr}"] = None

# write
out_row = 3
for r in rows:
    if r[0] == "__SECTION__":
        label = r[1]
        ws.merge_cells(start_row=out_row, start_column=1, end_row=out_row, end_column=len(COLS))
        c = ws.cell(out_row, 1, label)
        # style like the column-header (bold, filled) but left-aligned
        style_like(c, hdr_style_cell, wrap=True, align_h="left", align_v="center")
        for j in range(2, len(COLS) + 1):
            cc = ws.cell(out_row, j)
            style_like(cc, hdr_style_cell, wrap=True, align_h="left", align_v="center")
        ws.row_dimensions[out_row].height = 18
        out_row += 1
        continue
    for j, val in enumerate(r):
        c = ws.cell(out_row, j + 1, val)
        style_like(c, data_style_cell, wrap=True,
                   align_h=("center" if COLS[j] in ("B", "E", "H", "I", "J", "K") else "left"),
                   align_v="center")
        # tint genuinely-open / tbd values lightly
    ws.row_dimensions[out_row].height = 28
    out_row += 1

# widen columns for readability (keep existing where set, widen the rest)
widths = {"A": 34, "B": 13, "C": 24, "D": 26, "E": 11, "F": 28, "G": 28, "H": 12, "I": 14, "J": 12, "K": 24}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

wb.save(OUT)
print("WROTE", OUT)
print("ZUG-LAND now has", out_row - 1, "rows (header + sections + flows)")
