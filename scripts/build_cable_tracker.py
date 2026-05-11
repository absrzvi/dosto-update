from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

wb = Workbook()

NAVY = "1F4E79"
MIDBLUE = "2E75B6"
LIGHTBLUE = "D6E4F0"
GREENBG = "E2EFDA"
REDBG = "FCE4D6"
YELLOWBG = "FFF2CC"
GREYBG = "F2F2F2"
BORDER_GREY = "BFBFBF"

thin = Side(style="thin", color=BORDER_GREY)
box = Border(top=thin, bottom=thin, left=thin, right=thin)

font_name = "Arial"

def set_default_font(ws):
    pass

# ── Sheet 1: Issues ────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Issues"

headers = [
    "ID", "Trainset", "Consist", "Switch / Link", "Port(s)",
    "Fault Type", "Observed (LLDP/Port state)", "Expected (per IP-allocation plan)",
    "Required Action", "Status", "Date Found", "Date Resolved",
    "Stadler Ref / Ticket", "Verified By", "Notes",
]

for col, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = Font(name=font_name, bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = box

ws.row_dimensions[1].height = 32
ws.freeze_panes = "A2"

widths = {
    "A": 6, "B": 12, "C": 9, "D": 16, "E": 16,
    "F": 18, "G": 36, "H": 36, "I": 42, "J": 12,
    "K": 13, "L": 13, "M": 18, "N": 14, "O": 30,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

rows = [
    [1, "4734-101", "4-car", "E2 ↔ B1", "E2.e0-0 ↔ B1.e0-1",
     "wrong neighbour",
     "E2.e0-0 reaches B1; B1.e0-1 reaches E2",
     "E2.e0-0 ↔ E3 (intra-coach); E↔B inter-coach is E1.e0-0 ↔ B1.e0-1",
     "Re-patch E-coach end so E2.e0-0 lands on E3, and inter-coach E↔B trunk lands on E1.e0-0 ↔ B1.e0-1.",
     "OPEN", "2026-05-06", "", "", "Nomad Digital", "Found via lldp_topology_check.py"],
    [2, "4736-108", "6-car", "C3", "e0-0 / e0-1",
     "cable swap",
     "C3.e0-0 → C2; C3.e0-1 → A2",
     "C3.e0-0 ↔ A2; C3.e0-1 ↔ C2",
     "Swap the two trunk cables on C3. After swap, 'show lldp neighbours' on C3 should show e0-0=A2, e0-1=C2.",
     "OPEN", "2026-05-04", "", "", "Nomad Digital",
     "See Stadler_4736-108_Cabling_Fault_Report_v1.0.docx"],
    [3, "4736-108", "6-car", "D1 ↔ E2", "D1.e0-1 / E2.e0-1",
     "missing trunk",
     "No LLDP peer either end; both switches reachable on management VLAN",
     "Inter-coach trunk D1.e0-1 ↔ E2.e0-1",
     "Locate and reconnect inter-coach trunk cable between D1 e0-1 and E2 e0-1. Install replacement if absent.",
     "OPEN", "2026-05-04", "", "", "Nomad Digital",
     "See Stadler_4736-108_Cabling_Fault_Report_v1.0.docx"],
    [4, "4736-109", "6-car", "B3", "e0-4 (AP trunk)",
     "AP not connected",
     "Port admin-enabled, link DOWN, PoE drawing 0 W, no traffic ever",
     "AP physically connected to B3 e0-4 with link UP at 1G full-duplex",
     "Verify whether an AP is physically installed at B3 position. If yes, connect the patch cable to B3 e0-4.",
     "OPEN", "2026-05-04", "", "", "Nomad Digital",
     "See Stadler_4736-109_L2_Health_Check_Report_v1.0.docx"],
]

for r_idx, row in enumerate(rows, start=2):
    for c_idx, val in enumerate(row, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.font = Font(name=font_name, size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True,
                                horizontal="center" if c_idx in (1, 3, 10, 11, 12) else "left")
        c.border = box
    ws.row_dimensions[r_idx].height = 60

last_row_with_data = ws.max_row
buffer_rows = 100
total_rows = last_row_with_data + buffer_rows
last_col = len(headers)
last_col_letter = get_column_letter(last_col)

# Pre-format the buffer rows so new entries inherit borders/wrap.
for r in range(last_row_with_data + 1, total_rows + 1):
    for c_idx in range(1, last_col + 1):
        c = ws.cell(row=r, column=c_idx, value=None)
        c.font = Font(name=font_name, size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True,
                                horizontal="center" if c_idx in (1, 3, 10, 11, 12) else "left")
        c.border = box
    ws.row_dimensions[r].height = 36

table_ref = f"A1:{last_col_letter}{total_rows}"
table = Table(displayName="CableIssues", ref=table_ref)
table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False,
)
ws.add_table(table)

# Data validation: Status, Fault Type, Consist
status_dv = DataValidation(
    type="list",
    formula1='"OPEN,IN PROGRESS,RESOLVED,WONTFIX"',
    allow_blank=True, showDropDown=False,
)
status_dv.error = "Pick one of: OPEN, IN PROGRESS, RESOLVED, WONTFIX"
status_dv.errorTitle = "Invalid status"
ws.add_data_validation(status_dv)
status_dv.add(f"J2:J{total_rows}")

fault_dv = DataValidation(
    type="list",
    formula1='"cable swap,wrong neighbour,wrong far-end port,missing trunk,AP not connected,faulty SFP/cable,other"',
    allow_blank=True, showDropDown=False,
)
fault_dv.error = "Pick a fault type from the list (or 'other')."
fault_dv.errorTitle = "Invalid fault type"
ws.add_data_validation(fault_dv)
fault_dv.add(f"F2:F{total_rows}")

consist_dv = DataValidation(
    type="list", formula1='"4-car,6-car"', allow_blank=True, showDropDown=False,
)
ws.add_data_validation(consist_dv)
consist_dv.add(f"C2:C{total_rows}")

# Conditional formatting on Status (column J)
green_fill = PatternFill("solid", start_color=GREENBG)
red_fill = PatternFill("solid", start_color=REDBG)
yellow_fill = PatternFill("solid", start_color=YELLOWBG)
grey_fill = PatternFill("solid", start_color=GREYBG)

ws.conditional_formatting.add(
    f"J2:J{total_rows}",
    CellIsRule(operator="equal", formula=['"OPEN"'], fill=red_fill,
               font=Font(name=font_name, size=10, bold=True, color="9C0006")),
)
ws.conditional_formatting.add(
    f"J2:J{total_rows}",
    CellIsRule(operator="equal", formula=['"IN PROGRESS"'], fill=yellow_fill,
               font=Font(name=font_name, size=10, bold=True, color="9C5700")),
)
ws.conditional_formatting.add(
    f"J2:J{total_rows}",
    CellIsRule(operator="equal", formula=['"RESOLVED"'], fill=green_fill,
               font=Font(name=font_name, size=10, bold=True, color="375623")),
)
ws.conditional_formatting.add(
    f"J2:J{total_rows}",
    CellIsRule(operator="equal", formula=['"WONTFIX"'], fill=grey_fill,
               font=Font(name=font_name, size=10, bold=True, color="595959")),
)

# Highlight whole row faintly when RESOLVED so eye skips them.
ws.conditional_formatting.add(
    f"A2:{last_col_letter}{total_rows}",
    FormulaRule(formula=[f'$J2="RESOLVED"'],
                fill=PatternFill("solid", start_color="EBF1DE")),
)

# Comments on key headers explaining what to put.
ws["A1"].comment = Comment("Sequential ID. Don't reuse retired IDs.", "Nomad Digital")
ws["D1"].comment = Comment(
    "Use generic switch IDs like A1, D2, E3, B1.\n"
    "For inter-coach links use 'X ↔ Y' (e.g. 'D1 ↔ E2').", "Nomad Digital")
ws["E1"].comment = Comment(
    "Port labels like e0-0, e0-1, e1-8, etc.\n"
    "For trunk pairs: 'X.e0-0 / Y.e0-1'.", "Nomad Digital")
ws["F1"].comment = Comment(
    "Pick from drop-down. Definitions:\n"
    "- cable swap: both cables present, plugged into wrong ports on the same switch.\n"
    "- wrong neighbour: cable goes to the wrong far-end switch.\n"
    "- wrong far-end port: correct switch, wrong port number.\n"
    "- missing trunk: no LLDP either end, cable absent or unplugged.\n"
    "- AP not connected: AP trunk admin-enabled, no link / no PoE draw.\n"
    "- faulty SFP/cable: link UP but error counters non-zero.",
    "Nomad Digital")
ws["J1"].comment = Comment(
    "OPEN: Stadler action pending.\n"
    "IN PROGRESS: Stadler working on it.\n"
    "RESOLVED: re-verified clean by Nomad.\n"
    "WONTFIX: accepted as-is (e.g. AP not installed by design).",
    "Nomad Digital")

# ── Sheet 2: Dashboard ────────────────────────────────────────────────────────
dash = wb.create_sheet("Dashboard")
dash.sheet_view.showGridLines = False

dash["B2"] = "DOSTO Cable Issues Tracker — Dashboard"
dash["B2"].font = Font(name=font_name, bold=True, size=18, color=NAVY)
dash["B3"] = "Snapshot of open vs. resolved cabling issues across the DOSTO fleet."
dash["B3"].font = Font(name=font_name, italic=True, size=10, color="595959")

dash["B5"] = "Status"
dash["C5"] = "Count"
for cell in (dash["B5"], dash["C5"]):
    cell.font = Font(name=font_name, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color=MIDBLUE)
    cell.alignment = Alignment(horizontal="center")
    cell.border = box

statuses = [("OPEN", REDBG), ("IN PROGRESS", YELLOWBG), ("RESOLVED", GREENBG), ("WONTFIX", GREYBG)]
for i, (label, fill_col) in enumerate(statuses):
    r = 6 + i
    dash.cell(row=r, column=2, value=label).font = Font(name=font_name, bold=True)
    dash.cell(row=r, column=2).fill = PatternFill("solid", start_color=fill_col)
    dash.cell(row=r, column=2).alignment = Alignment(horizontal="left", indent=1)
    dash.cell(row=r, column=2).border = box
    formula = f'=COUNTIF(Issues!J:J,"{label}")'
    dash.cell(row=r, column=3, value=formula).font = Font(name=font_name, bold=True, size=12)
    dash.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    dash.cell(row=r, column=3).border = box

dash["B10"] = "TOTAL"
dash["B10"].font = Font(name=font_name, bold=True, color="FFFFFF")
dash["B10"].fill = PatternFill("solid", start_color=NAVY)
dash["B10"].alignment = Alignment(horizontal="left", indent=1)
dash["B10"].border = box
dash["C10"] = "=SUM(C6:C9)"
dash["C10"].font = Font(name=font_name, bold=True, color="FFFFFF", size=12)
dash["C10"].fill = PatternFill("solid", start_color=NAVY)
dash["C10"].alignment = Alignment(horizontal="center")
dash["C10"].border = box

dash["E5"] = "Open by Trainset"
dash["E5"].font = Font(name=font_name, bold=True, color="FFFFFF")
dash["E5"].fill = PatternFill("solid", start_color=MIDBLUE)
dash["E5"].alignment = Alignment(horizontal="center")
dash["E5"].border = box
dash["F5"] = "Open"
dash["F5"].font = Font(name=font_name, bold=True, color="FFFFFF")
dash["F5"].fill = PatternFill("solid", start_color=MIDBLUE)
dash["F5"].alignment = Alignment(horizontal="center")
dash["F5"].border = box

trainsets_seen = sorted({r[1] for r in rows})
for i, ts in enumerate(trainsets_seen):
    r = 6 + i
    dash.cell(row=r, column=5, value=ts).font = Font(name=font_name)
    dash.cell(row=r, column=5).alignment = Alignment(horizontal="center")
    dash.cell(row=r, column=5).border = box
    dash.cell(row=r, column=6,
              value=f'=COUNTIFS(Issues!B:B,"{ts}",Issues!J:J,"OPEN")')
    dash.cell(row=r, column=6).font = Font(name=font_name, bold=True)
    dash.cell(row=r, column=6).alignment = Alignment(horizontal="center")
    dash.cell(row=r, column=6).border = box

dash["E12"] = "Open by Fault Type"
dash["E12"].font = Font(name=font_name, bold=True, color="FFFFFF")
dash["E12"].fill = PatternFill("solid", start_color=MIDBLUE)
dash["E12"].alignment = Alignment(horizontal="center")
dash["E12"].border = box
dash["F12"] = "Open"
dash["F12"].font = Font(name=font_name, bold=True, color="FFFFFF")
dash["F12"].fill = PatternFill("solid", start_color=MIDBLUE)
dash["F12"].alignment = Alignment(horizontal="center")
dash["F12"].border = box

fault_types = ["cable swap", "wrong neighbour", "wrong far-end port",
               "missing trunk", "AP not connected", "faulty SFP/cable", "other"]
for i, ft in enumerate(fault_types):
    r = 13 + i
    dash.cell(row=r, column=5, value=ft).font = Font(name=font_name)
    dash.cell(row=r, column=5).alignment = Alignment(horizontal="left", indent=1)
    dash.cell(row=r, column=5).border = box
    dash.cell(row=r, column=6,
              value=f'=COUNTIFS(Issues!F:F,"{ft}",Issues!J:J,"OPEN")')
    dash.cell(row=r, column=6).font = Font(name=font_name, bold=True)
    dash.cell(row=r, column=6).alignment = Alignment(horizontal="center")
    dash.cell(row=r, column=6).border = box

dash["B22"] = "How to use this tracker"
dash["B22"].font = Font(name=font_name, bold=True, size=12, color=NAVY)
notes = [
    "1. Add new findings as rows on the Issues sheet (next blank row in the table).",
    "2. Use generic switch IDs only (A1…F3 / G1…G3) — no IPs, no MACs, no live hostnames.",
    "3. Fault Type, Status, and Consist have drop-downs — pick from the list.",
    "4. When Stadler fixes an issue: re-run lldp_topology_check.py, then set Status = RESOLVED and fill Date Resolved.",
    "5. Counts on this dashboard refresh automatically.",
]
for i, n in enumerate(notes):
    dash.cell(row=23 + i, column=2, value=n).font = Font(name=font_name, size=10)

dash.column_dimensions["A"].width = 2
dash.column_dimensions["B"].width = 22
dash.column_dimensions["C"].width = 12
dash.column_dimensions["D"].width = 2
dash.column_dimensions["E"].width = 24
dash.column_dimensions["F"].width = 10

# ── Sheet 3: Fault Type Reference ─────────────────────────────────────────────
ref = wb.create_sheet("Fault Types")
ref.sheet_view.showGridLines = False
ref["B2"] = "Fault Type Reference"
ref["B2"].font = Font(name=font_name, bold=True, size=16, color=NAVY)

ref_headers = ["Fault Type", "Definition", "How we detect it", "Stadler vs Nomad"]
for c_idx, h in enumerate(ref_headers, start=2):
    cell = ref.cell(row=4, column=c_idx, value=h)
    cell.font = Font(name=font_name, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = box

ref_rows = [
    ["cable swap",
     "Both inter-switch trunk cables present, but plugged into wrong ports on the same switch (e0-0 and e0-1 transposed).",
     "lldp_topology_check.py reports two MISMATCHes on the same switch where the live peers are each other's expected peer.",
     "Stadler — re-patch."],
    ["wrong neighbour",
     "Cable goes to the wrong far-end switch (e.g. cable that should reach E1 is plugged into E2 instead).",
     "lldp_topology_check.py: live peer ≠ expected peer, and the live peer is a different switch in the same coach group.",
     "Stadler — re-patch."],
    ["wrong far-end port",
     "Cable lands on the correct far-end switch, but on the wrong port number.",
     "Inspect both ends; expected port shows NO NEIGHBOUR, an unexpected port shows the link.",
     "Stadler — re-patch."],
    ["missing trunk",
     "Inter-coach trunk has no LLDP either end. Cable absent, unplugged, or routed into a dead port.",
     "Both expected endpoints show NO NEIGHBOUR on the affected port. Both switches reachable on management VLAN.",
     "Stadler — locate / install."],
    ["AP not connected",
     "AP trunk port admin-enabled, link DOWN, PoE drawing 0 W, no traffic ever.",
     "show interface summary: DOWN. show interface details: 0 RX/TX bytes. PoE: 0 W draw.",
     "Stadler — install AP or patch cable."],
    ["faulty SFP/cable",
     "Link UP but per-port error counters are non-zero (RX CRC, carrier-false, etc.) above noise level.",
     "show interface <port> details: RX crc errors, carrier false, or pause frames > 0 sustained.",
     "Stadler — replace SFP or cable."],
    ["other",
     "Anything that doesn't fit the categories above.",
     "Document fully in the Notes column on the Issues sheet.",
     "Decide per case."],
]

for r_idx, row in enumerate(ref_rows, start=5):
    for c_idx, val in enumerate(row, start=2):
        c = ref.cell(row=r_idx, column=c_idx, value=val)
        c.font = Font(name=font_name, size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
        c.border = box
    ref.row_dimensions[r_idx].height = 60

ref.column_dimensions["A"].width = 2
ref.column_dimensions["B"].width = 22
ref.column_dimensions["C"].width = 50
ref.column_dimensions["D"].width = 50
ref.column_dimensions["E"].width = 24

ref["B14"] = "Note: Template / OBN-config issues (duplicate hostnames, factory dosto-00000000 switches, missing nv4/nv6 templates) are NOT cable issues. They belong in the OBN troubleshooting log, not in this tracker."
ref["B14"].font = Font(name=font_name, italic=True, size=10, color="C00000")
ref.merge_cells("B14:E14")
ref["B14"].alignment = Alignment(wrap_text=True, vertical="top")

wb.active = wb.sheetnames.index("Dashboard")

out = r"C:\Users\AbbasRizvi\Documents\dosto-troubleshooting\cable-issues-tracker.xlsx"
wb.save(out)
print(f"Saved: {out}")
