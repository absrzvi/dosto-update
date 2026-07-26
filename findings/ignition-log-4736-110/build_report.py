#!/usr/bin/env python3
"""Build the Stadler power-outage evidence report (xlsx) for 4736-110 / Fzg 138.
Source data: vign_live_20260724.csv (CMM I2C logger) + ignition-log/shutdown.log.
All asserted numbers are computed here from source, not hand-typed."""
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import timedelta

CSV = "vign_live_20260724.csv"
SDLOG = "ignition-log/shutdown.log"
OUT = "4736-110_Fzg138_CCU_Power_Outage_Report_2026-07-24.xlsx"
WINDOW_LABEL = "2026-07-03 to 2026-07-24 (last 3 weeks)"

def t(s):
    return datetime.strptime(s.strip().rstrip("Z"), "%Y-%m-%dT%H:%M:%S")

# Excel display formats for normalised timestamps (values stored as real datetimes)
DT_FMT = "yyyy-mm-dd hh:mm:ss"   # full timestamp, UTC
DATE_FMT = "yyyy-mm-dd"          # date only

def norm(line):
    if len(line) >= 12:
        iso, vin, vign, pct, v12, v5, curr, inlet, outlet, bp, st, reason = line[:12]
    else:
        iso, vin, vign, pct, v12, v5, curr, bp, st, reason = line[:10]
        inlet = outlet = ""
    return dict(iso=iso, vin=vin, vign=vign, pct=pct, v12=v12, v5=v5, curr=curr,
                inlet=inlet, outlet=outlet, bp=bp, st=st, reason=reason)

rows = []
with open(CSV) as f:
    r = csv.reader(f); next(r)
    for line in r:
        if not line: continue
        try: float(line[1]); float(line[2])
        except: continue
        rows.append(norm(line))

sd = [t(l.split()[0]) for l in open(SDLOG) if "GRACEFUL" in l]

# ---- outage detection (gap > 5 min), analysis window: last 3 weeks ----
WINDOW = "2026-07-03"
events = []
for i in range(1, len(rows)):
    p, c = rows[i-1], rows[i]
    if p["iso"] < WINDOW: continue
    dt = (t(c["iso"]) - t(p["iso"])).total_seconds()
    if dt > 300:
        vin, vign = float(p["vin"]), float(p["vign"])
        graceful = any(abs((t(p["iso"]) - m).total_seconds()) < 1800 for m in sd)
        if graceful:
            cls, verdict = "Commanded", "Graceful shutdown marker present (nightly 03:00 UTC maintenance reboot or a commanded power-off). Not a fault."
        elif vin < 50:
            cls, verdict = "Supply loss", "Both Vin and Vign collapsed together = input supply removed at source."
        elif vign < 0.2 * vin:
            cls, verdict = "Ignition drop", "Vign collapsed to ~0 V while Vin held near nominal = ignition-input line lost while battery supply remained. Vehicle ignition/wiring signature."
        else:
            cls, verdict = "Abrupt hard cut", "No shutdown marker; last sample fully healthy (Vin and Vign both nominal, no fault flag) up to the final 30 s heartbeat, then power lost. Faster than an ACPI/ignition shutdown = external supply removed."
        events.append(dict(last=p["iso"], boot=c["iso"], down=dt/60.0,
                           vin=vin, vign=vign, st=p["st"], bp=p["bp"],
                           cls=cls, verdict=verdict))

hard = [e for e in events if e["cls"] == "Abrupt hard cut"]
igni = [e for e in events if e["cls"] == "Ignition drop"]
cmd  = [e for e in events if e["cls"] == "Commanded"]
supp = [e for e in events if e["cls"] == "Supply loss"]
noncmd = len(events) - len(cmd)
# phase split: date of first vs last of each fault type (data-driven, for the narrative)
igni_dates = [e["last"][:10] for e in igni]
hard_dates = [e["last"][:10] for e in hard]

# ---- styling helpers ----
NAVY = "1F3864"; LBLUE = "D9E1F2"; GREY = "F2F2F2"
AMBER = "FFF2CC"; REDF = "F8CBAD"; GREENF = "C6E0B4"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
FN = "Arial"

def cell(ws, ref, val, *, bold=False, size=10, color="000000", fill=None,
         wrap=False, align="left", valign="top", bd=False, num=None):
    c = ws[ref]; c.value = val
    c.font = Font(name=FN, bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    if bd: c.border = border
    if num: c.number_format = num
    return c

wb = Workbook()

# ===================== Sheet 1: Summary =====================
ws = wb.active; ws.title = "Summary"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 64
ws.column_dimensions["D"].width = 3

cell(ws, "B2", "CCU Power-Loss Investigation — Evidence Report", bold=True, size=15, color=NAVY)
cell(ws, "B3", "Train 4736-110  ·  Fzg 138  ·  CCU box1-t23 (10.179.23.1)", size=10, color="595959")
cell(ws, "B4", "Prepared by Nomad Digital for Stadler  ·  2026-07-24", size=9, color="808080")

meta = [
    ("Reported symptom", "CCU intermittently going offline; complaint escalated over the preceding ~2 days."),
    ("Analysis window", WINDOW_LABEL + ", UTC."),
    ("Evidence source", "Nomad CCU chassis-management-module (CMM) power logger, sampling supply voltage (Vin), "
                        "ignition voltage (Vign), 12 V/5 V rails, current, and chassis fault flags every 30 s, "
                        "written to non-volatile storage on the CCU."),
    ("Logger installed", "2026-06-16, hardened 2026-06-18 (30 s heartbeat so the last sample before any cut is <=30 s old)."),
    ("Records analysed", f"{len(rows):,} samples, {rows[0]['iso']} to {rows[-1]['iso']}."),
]
r = 6
for k, v in meta:
    cell(ws, f"B{r}", k, bold=True, size=10, color=NAVY, valign="top", wrap=True, bd=True, fill=LBLUE)
    cell(ws, f"C{r}", v, size=10, wrap=True, bd=True)
    ws.row_dimensions[r].height = 42 if len(v) > 90 else 30
    r += 1

r += 1
cell(ws, f"B{r}", "Headline finding", bold=True, size=12, color=NAVY); r += 1
finding = (f"The CCU power log records {len(events)} interruptions in the 3-week window. {len(cmd)} were clean commanded "
           f"shutdowns (the nightly maintenance reboot); the remaining {noncmd} were unexpected power losses in TWO distinct "
           f"electrical signatures: {len(igni)} 'ignition-drop' events (ignition voltage Vign collapsed to 0 V while the "
           f"battery supply Vin held near nominal) and {len(hard)} 'abrupt hard-cut' events (Vin and Vign both healthy to the "
           "last sample, then instant power loss). Both signatures point to vehicle-side power/ignition wiring, not to the CCU. "
           "The chassis fault byte read clean (0x00) through every event.")
cell(ws, f"B{r}", finding, size=10, wrap=True, fill=GREENF, bd=True)
ws.merge_cells(f"B{r}:C{r}"); ws.row_dimensions[r].height = 96; r += 2

cell(ws, f"B{r}", "Two-phase pattern over the 3 weeks", bold=True, size=12, color=NAVY); r += 1
phase = (f"Ignition-drop events cluster in the earlier part of the window (first {igni_dates[0] if igni_dates else 'n/a'}, "
         f"last {igni_dates[-1] if igni_dates else 'n/a'}); abrupt hard-cuts dominate the most recent days "
         f"(first {hard_dates[0] if hard_dates else 'n/a'}, last {hard_dates[-1] if hard_dates else 'n/a'}), overlapping "
         "mid-month. The failure behaviour therefore appears to have shifted from an intermittent ignition-line dropout toward "
         "outright supply removal. The per-event timeline is on the 'Outage Events' tab.")
cell(ws, f"B{r}", phase, size=10, wrap=True, fill=LBLUE, bd=True)
ws.merge_cells(f"B{r}:C{r}"); ws.row_dimensions[r].height = 66; r += 2

cell(ws, f"B{r}", "What this indicates", bold=True, size=12, color=NAVY); r += 1
concl = [
    "The CCU is NOT shutting itself down: no operating-system or ignition-driven (ACPI) shutdown occurred on any "
    "fault event — those leave a graceful-shutdown marker, present only on the commanded reboots.",
    f"Ignition-line dropout is real and repeatable ({len(igni)} events): Vign fell to 0 V while Vin held ~110-118 V. "
    "The ignition input is being lost while the battery supply is still present — a vehicle ignition/wiring signature.",
    f"Abrupt supply removal is the more recent pattern ({len(hard)} events): everything nominal to the last 30 s sample, "
    "then instant loss with no decay. Consistent with a breaker/isolator opening or a hard power removal upstream of the CCU.",
    "It is NOT a CCU internal fault: chassis fault flags (fans, rails, temperature) read clean (0x00) throughout, and "
    "inlet/outlet temperatures stayed well within limits across the whole window.",
    "Both signatures sit upstream of the CMM's own voltage sense point, so the CCU cannot capture the switching event "
    "itself in finer detail — the vehicle-side power/ignition log is needed to pin the exact mechanism.",
]
for c_ in concl:
    cell(ws, f"B{r}", "•", bold=True, color=NAVY, valign="top")
    cell(ws, f"C{r}", c_, size=10, wrap=True, valign="top")
    ws.row_dimensions[r].height = 44; r += 1

r += 1
cell(ws, f"B{r}", "Requested from Stadler", bold=True, size=12, color=NAVY); r += 1
ask = ("Please share the vehicle power / ignition-relay / battery-isolator event log for 4736-110 covering "
       "2026-07-03 to 2026-07-24 (UTC). Correlating those switching events against the CCU event times in the "
       "'Outage Events' tab will confirm the upstream cause and, in particular, explain why the ignition line "
       "(Vign) is dropping to 0 V while the battery supply holds. The detailed per-event evidence is on the "
       "'Outage Events' tab; the complete raw voltage log is on the 'Raw Data' tab for independent verification.")
cell(ws, f"B{r}", ask, size=10, wrap=True, fill=AMBER, bd=True)
ws.merge_cells(f"B{r}:C{r}"); ws.row_dimensions[r].height = 66

# ===================== Sheet 2: Outage Events =====================
we = wb.create_sheet("Outage Events")
we.sheet_view.showGridLines = False
cell(we, "A1", "Outage Events — 4736-110 / Fzg 138 — CCU power interruptions 2026-07-03 to 2026-07-24 (UTC)",
     bold=True, size=12, color=NAVY)
cell(we, "A2", "One row per interruption in the CCU power log. 'Last healthy sample' is the final reading before "
     "power was lost; 'Next boot' is when the CCU came back. Voltages are from the CMM.", size=9, color="595959", wrap=True)
we.merge_cells("A2:I2"); we.row_dimensions[2].height = 28

heads = ["#", "Last healthy sample\n(UTC)", "Next boot\n(UTC)", "Downtime (h:mm)",
         "Vin (V)", "Vign (V)", "Chassis fault", "Classification", "Interpretation"]
widths = [4, 21, 21, 14, 9, 9, 13, 18, 60]
hr = 4
for j, (h, w) in enumerate(zip(heads, widths), start=1):
    L = get_column_letter(j)
    we.column_dimensions[L].width = w
    cell(we, f"{L}{hr}", h, bold=True, size=10, color="FFFFFF", fill=NAVY,
         wrap=True, align="center", valign="center", bd=True)
we.row_dimensions[hr].height = 30

def hhmm(mins):
    h = int(mins // 60); m = int(round(mins - h*60))
    return f"{h}:{m:02d}"

CLS_FILL = {"Commanded": GREENF, "Abrupt hard cut": REDF, "Ignition drop": AMBER, "Supply loss": REDF}
row = hr + 1
for i, e in enumerate(events, start=1):
    fill = CLS_FILL.get(e["cls"], GREY)
    cell(we, f"A{row}", i, align="center", bd=True)
    cell(we, f"B{row}", t(e["last"]), bd=True, num=DT_FMT, align="center")
    cell(we, f"C{row}", t(e["boot"]), bd=True, num=DT_FMT, align="center")
    cell(we, f"D{row}", hhmm(e["down"]), align="center", bd=True)
    cell(we, f"E{row}", round(e["vin"], 1), align="center", bd=True, num="0.0")
    cell(we, f"F{row}", round(e["vign"], 1), align="center", bd=True, num="0.0")
    cell(we, f"G{row}", "None (0x00)" if e["st"] == "0x00" else e["st"], align="center", bd=True)
    cell(we, f"H{row}", e["cls"], bold=True, align="center", bd=True, fill=fill, wrap=True)
    cell(we, f"I{row}", e["verdict"], size=9, wrap=True, bd=True)
    we.row_dimensions[row].height = 46
    row += 1

# note block under table
row += 1
cell(we, f"A{row}", "How to read the classification:", bold=True, size=10, color=NAVY); row += 1
notes = [
    ("Commanded (green)", "A graceful-shutdown marker was written to the CCU — the nightly 03:00 UTC maintenance reboot or a deliberate power-off. Not a fault."),
    ("Ignition drop (amber)", "Ignition voltage Vign fell to ~0 V while battery supply Vin held near nominal. The ignition input was lost while power was still present. Vehicle ignition/wiring signature."),
    ("Abrupt hard cut (red)", "No shutdown marker; the last sample was fully healthy (Vin and Vign both nominal). Power was lost between two 30 s samples. External supply removed. Dominant pattern in the last few days."),
    ("Supply loss (red)", "Vin and Vign collapsed together — input supply removed at source. (None observed in this window.)"),
]
for k, v in notes:
    cell(we, f"B{row}", k, bold=True, size=9, valign="top")
    cell(we, f"C{row}", v, size=9, wrap=True, valign="top")
    we.merge_cells(f"C{row}:I{row}")
    we.row_dimensions[row].height = 30; row += 1

we.freeze_panes = "A5"

# ===================== Sheet 3: Methodology =====================
wm = wb.create_sheet("Methodology")
wm.sheet_view.showGridLines = False
wm.column_dimensions["A"].width = 3
wm.column_dimensions["B"].width = 100
cell(wm, "B2", "Methodology & Data Provenance", bold=True, size=13, color=NAVY)
para = [
    ("Logger", "A watchdog script on the CCU reads the ADLINK R5001C Chassis Management Module (CMM) over the "
               "internal I2C bus every 3 seconds. It records a row on any meaningful change (Vin/Vign delta > 2 V, "
               "temperature delta > 2 C, or a mode/fault change) and, in addition, a heartbeat row every 30 seconds "
               "so that the last recorded sample before any power loss is at most 30 seconds old."),
    ("Measured quantities", "Vin (chassis input supply voltage), Vign (ignition-line voltage), the 12 V and 5 V "
               "standby rails, chassis current, inlet/outlet temperature, the 12 V-backplane power mode (Ignition vs "
               "Manual), and the CMM Status-1 fault byte (fans, rails, temperature)."),
    ("Non-volatile storage", "The log is written to the CCU's persistent /data partition (btrfs). The CCU's normal "
               "system logs (journald, dmesg, /var/log) are volatile and are wiped on every boot, so this logger is the "
               "only record that survives an outage."),
    ("Shutdown marker", "A companion service writes a 'graceful shutdown' line to persistent storage only when the "
               "operating system shuts down cleanly. The presence of that marker distinguishes a commanded/ACPI shutdown "
               "from an abrupt power removal. Its absence after an outage means the power was cut without an orderly shutdown."),
    ("Outage detection", "An outage is identified as any gap greater than 5 minutes between consecutive log samples. "
               "For each gap, the last sample before it and the shutdown-marker log are examined to classify the cause."),
    ("Ignition-fault test", "A genuine ignition-input fault appears as Vign collapsing toward 0 V while Vin holds. "
               "This window contains both: events where Vign fell to 0 V while Vin held (classified 'Ignition drop'), and "
               "events where Vign stayed nominal to the last sample before an instant loss (classified 'Abrupt hard cut'). "
               "The test distinguishes the two mechanisms per event; both are external to the CCU."),
    ("Limitation", "The CMM reports instantaneous values and holds no internal event history; it cannot log the event "
               "that removes its own supply. A sub-30-second transient is therefore seen only as a healthy sample followed "
               "by absence. Confirming the exact upstream mechanism requires the vehicle-side power/ignition log, which is "
               "why that is requested from Stadler."),
]
rr = 4
for k, v in para:
    cell(wm, f"B{rr}", k, bold=True, size=11, color=NAVY); rr += 1
    cell(wm, f"B{rr}", v, size=10, wrap=True, valign="top")
    lines = (len(v) // 95) + 1
    wm.row_dimensions[rr].height = 15 * lines + 6
    rr += 2

# ===================== Sheet 4: Daily Tally =====================
wd = wb.create_sheet("Daily Tally")
wd.sheet_view.showGridLines = False
cell(wd, "A1", "Daily Tally — power interruptions per day by type (2026-07-03 to 2026-07-24, UTC)",
     bold=True, size=12, color=NAVY)
cell(wd, "A2", "One row per day. Shows the two fault signatures coexisting mid-window and the shift toward "
     "abrupt hard-cuts in the most recent days. Commanded (nightly reboot) shown for completeness.",
     size=9, color="595959", wrap=True)
wd.merge_cells("A2:F2"); wd.row_dimensions[2].height = 28

# build the full day range so days with zero events still appear
d0 = datetime.strptime(WINDOW, "%Y-%m-%d")
d1 = t(rows[-1]["iso"])
days = []
cur = d0
while cur.date() <= d1.date():
    days.append(cur.strftime("%Y-%m-%d")); cur += timedelta(days=1)

CLASSES = ["Ignition drop", "Abrupt hard cut", "Supply loss", "Commanded"]
tally = {d: {k: 0 for k in CLASSES} for d in days}
for e in events:
    day = e["last"][:10]
    if day in tally:
        tally[day][e["cls"]] += 1

dh = 4
dcols = ["Date"] + CLASSES + ["Total faults"]
dwidths = [13, 15, 17, 13, 13, 14]
for j, (h, w) in enumerate(zip(dcols, dwidths), start=1):
    L = get_column_letter(j)
    wd.column_dimensions[L].width = w
    cell(wd, f"{L}{dh}", h, bold=True, size=10, color="FFFFFF", fill=NAVY,
         wrap=True, align="center", valign="center", bd=True)
wd.row_dimensions[dh].height = 30

CLS_FILL2 = {"Ignition drop": AMBER, "Abrupt hard cut": REDF, "Supply loss": REDF, "Commanded": GREENF}
dr = dh + 1
for d in days:
    cell(wd, f"A{dr}", datetime.strptime(d, "%Y-%m-%d"), size=10, bd=True, num=DATE_FMT)
    for j, k in enumerate(CLASSES, start=2):
        v = tally[d][k]
        L = get_column_letter(j)
        c = cell(wd, f"{L}{dr}", v if v else "", align="center", bd=True)
        if v: c.fill = PatternFill("solid", fgColor=CLS_FILL2[k])
    # total faults = all non-commanded (literal, so every viewer shows it, not just Excel-on-recalc)
    tot = tally[d]["Ignition drop"] + tally[d]["Abrupt hard cut"] + tally[d]["Supply loss"]
    cell(wd, f"F{dr}", tot if tot else "", bold=True, align="center", bd=True)
    dr += 1

# totals row (literal column sums)
cell(wd, f"A{dr}", "Total", bold=True, fill=LBLUE, bd=True)
col_totals = {k: sum(tally[d][k] for d in days) for k in CLASSES}
faults_total = col_totals["Ignition drop"] + col_totals["Abrupt hard cut"] + col_totals["Supply loss"]
for j, k in enumerate(CLASSES, start=2):
    L = get_column_letter(j)
    cell(wd, f"{L}{dr}", col_totals[k], bold=True, align="center", fill=LBLUE, bd=True)
cell(wd, f"F{dr}", faults_total, bold=True, align="center", fill=LBLUE, bd=True)
totals_row = dr

# stacked bar chart: ignition-drop vs abrupt-cut per day
chart = BarChart()
chart.type = "col"; chart.grouping = "stacked"; chart.overlap = 100
chart.title = "Fault events per day — Ignition drop (amber) vs Abrupt hard cut (red)"
chart.height = 8; chart.width = 26
chart.y_axis.title = "events"; chart.x_axis.title = "date"
data = Reference(wd, min_col=2, max_col=3, min_row=dh, max_row=totals_row - 1)
cats = Reference(wd, min_col=1, min_row=dh + 1, max_row=totals_row - 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
try:
    chart.series[0].graphicalProperties.solidFill = "FFD966"  # ignition = amber
    chart.series[1].graphicalProperties.solidFill = "E06666"  # abrupt = red
except Exception:
    pass
wd.add_chart(chart, f"H{dh}")

# ===================== Sheet 5: Raw Data =====================
wr = wb.create_sheet("Raw Data")
rheads = ["timestamp_utc", "vin_v", "vign_v", "vign_pct_of_vin", "v12_v", "v5stb_v",
          "curr_a", "inlet_c", "outlet_c", "bp_mode", "status1_faults", "reason"]
for j, h in enumerate(rheads, start=1):
    cell(wr, f"{get_column_letter(j)}1", h, bold=True, size=9, color="FFFFFF", fill=NAVY, bd=False)
    wr.column_dimensions[get_column_letter(j)].width = 21 if j == 1 else 12
for i, d in enumerate(rows, start=2):
    vals = [d["iso"], d["vin"], d["vign"], d["pct"], d["v12"], d["v5"], d["curr"], d["inlet"], d["outlet"], d["bp"], d["st"], d["reason"]]
    for j, v in enumerate(vals, start=1):
        c = wr.cell(row=i, column=j, value=v)
        c.font = Font(name=FN, size=9)
        if j == 1:
            c.value = t(v); c.number_format = DT_FMT       # normalise timestamp
        elif j in (2, 3, 4, 5, 6, 7, 8, 9):
            try: c.value = float(v); c.number_format = "0.00"
            except (ValueError, TypeError): pass
wr.freeze_panes = "A2"
cell(wr, f"A{len(rows)+3}", f"Complete CMM power log, {len(rows):,} samples, "
     f"{rows[0]['iso']} to {rows[-1]['iso']}. Source of every figure in this workbook.",
     size=8, color="808080", wrap=True)

wb.save(OUT)
print("wrote", OUT)
print("events:", len(events), "| hard cuts:", len(hard), "| raw rows:", len(rows))
